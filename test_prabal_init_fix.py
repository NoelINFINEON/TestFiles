# To read more lines in VS code terminal -> "CTRL"+""," -> search for -> terminal.integrated.scrollback
# from asyncore import read
#To Install Packagee, use the command ---> pip install packageName
#*******************MODULE IMPORTS*******************************
from gzip import READ
import os,time,pytest,pylink,pyvisa,re,itertools,sys,time,openpyxl,collections
import pandas as pd
from PyPDF2 import PdfFileMerger
from time import perf_counter
# from pandas import DataFrame, ExcelWriter
from os import write
from unicodedata import name
from copy import copy
from typing import Union, Optional
import numpy as np
from openpyxl.utils import get_column_letter
from bitstring import Bits
# from test_excel import *
from alive_progress import alive_bar
import collections
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
from openpyxl.chart import Reference, LineChart , Series, BubbleChart
from datetime import datetime

from registerMatch import * #imports from another file
from smu_commands import *
# from test_prabal import *
from importRegMap import *
from power_supply_commands import *
from read_register import read_register_data
from power_cycle_init import full_power_cycle, power_cycle
#***********************************************************************


# from start_jlink import debugger 
# db = debugger()

# running_report = {'0':[],'1':[]} # 0=> not serious, 1=> serious and we need to stop testing
from start_helper_functions import helper_functions
db = helper_functions()
# from test_prabal_oven import *

########### Variable Initialization Here ###########
jlink_ENABLE = 1
NANOSECONDS_TO_SECONDS_CONVERSION_FACTOR = 1000000000
MILLISECONDS_TO_SECONDS_CONVERSION_FACTOR = 1000
stringName =''

sp_dict,ts_dict,is_dict,pi_dict,is_accuracy_dict ={},{},{},{},{}
saveMask ={}
dict = {}
newSave ={}
dataReadArray = [] #added this empty array for register read excel sheet
dataRegisterName = [] #added this empty array for register read excel sheet 
regcapture =[]

IS_LR_trim,VDDLO_trim,VDDA_trim,VDDLR_trim,VDDSCR_trim = [],[],[],[],[]
IS_LR_BANK,VDDLO_BANK,VDDLR_BANK,VDDA_BANK,VDDSCR_BANK = [],[],[],[],[]
BG_Target=0
VDDLO_Target=0
VDDLR_Target=0
VDDA_Target=0
IS_LR_Target = 0
VDDSCR_TARGET=0
TS_current_target=0
VDDLO_FLAG = 0
IS_LR_FLAG = 0
VDDLR_FLAG = 0
SP_BLOCK_FLAG = 0
IS_LR_BLOCK_FLAG = 0
VDDA_FLAG = 0
VDDSCR_FLAG = 0
LO_CLOCK_FLAG = 0
VGPTAT_FLAG = 0
BG_FLAG = 0
TS_CURR_TRIM_FLAG = 0
write_calculated_flag = 0
maskBits = 0
special_script = 0
readCommentReg = 0
readCounter = 0
counter = 0
excelDic = {}
calcdict = {}
exception_dict = {}
all_stored_vs = []
vs_adc_dict = {}
vs_adc_x = []
all_stored_vs_Tests = []

VS_Gain_Trim = {}
VS_Digital_Offset_BIST = {}
VS_Analog_Offset_BIST ={}
VS_Gain_Trim_Retest = {}
VS_Offset_BIST_retest = {}
VOS_DAC_offset_DNL_test = {}
DNL_Test = {}
pfd_folder = ''
this_file_name = ''

all_pfd = []
pfd_name = []
target_global = {}

expected_results = {}
looking_for_expected_ranges = 0
look_for_subsections = 0



debuging =1
debug_option = {}
    
#variableNameofEquip = pyvisa.ResourceManager().open_resource('GPIB0::(put last two digits of address)::INSTR')
SmuKeysight =  pyvisa.ResourceManager().open_resource('GPIB0::24::INSTR') #connect to the SMU
powerSupply =  pyvisa.ResourceManager().open_resource('GPIB0::08::INSTR') #connect to Power Supply 
dataAcq     =  pyvisa.ResourceManager().open_resource('GPIB0::10::INSTR') #connect to the data Acq Unit 
# OVEN_CONTROL     =  pyvisa.ResourceManager().open_resource('GPIB0::04::INSTR') #connect to the data Acq Unit - WORKING IN LAB 
# FUNC_GEN =  pyvisa.ResourceManager().open_resource('GPIB0::9::INSTR') #connect to the SMU




#Setting DAQ Channels that will be connected to the board
#Ch1 = TmuxN, Ch2 = TmuxP, Ch3 = VD12


#***********************BASIC BLOCK TRIM REGISTER NAMES******************************
#NOTE: Register Names Need To Be Updated For Different Projects
#Current Register Names Below: LANGLEY
#BG Block
name_BG = "trim.BG_TRIM1.bg_trim"
name_VGPTAT = "trim.BG_TRIM1.bg_vptat_trim"
#SP Block
name_VDDLO = "trim.BG_TRIM2.sp_reg_lo_trim"
name_VDDLR = "trim.BG_TRIM2.sp_reg_vdd_lr_trim"
name_VDDSCR = "trim.BG_TRIM2.sp_reg_vdd_scr_trim"
name_VDDA = "trim.BG_TRIM2.sp_rega_trim"
name_IS_LR = "trim.IS_VDD_LR_TRIM.is_vdd_lr_trim"
#LO BLock 
name_lo_clk_trim_coarse = "trim.LO_TRIM.lo_clk_trim_coarse"
name_lo_clk_trim_fine = "trim.LO_TRIM.lo_clk_trim_fine"
#************************************************************************************

#convert_ate_string_to_int will replace L's with 0's and H's with 1's in register read checks

#Register Read line that contain 'L' and 'H' do not perform anything during parsing
#Therefore, this funcation has no purpose at the moment
def convert_ate_string_to_int(arg):
    if 'L' in arg or 'H':
        arg = arg.replace('L', '0')  
        arg = arg.replace('H', '1')
    return arg

#maskMAN will simlarly replae X's,C's, and H's for register reads/masking purposes. Not being used at the moment
def maskMAN(arg):
    global maskBits
    arg = arg.replace('X', '0')
    arg = arg.replace('C', '1')
    arg = arg.replace('H', '1')
    """
    arg = int(arg, 2)
    maskBits = arg
    maskBits = bin(maskBits)[2:] #turning into binary
    maskBits = int(maskBits)
    """
    # db.printing(debugging,maskBits)
    return arg

def clean_swd_args(arg_list,special_script):
    # Removes any special characters or SWD tags
    # Converts from string to int
    # Converts ate string (LH) to binary
    
    characters_to_be_removed = "<>"
    cleaned_arg_list = []
    # db.printing(debugging,"sss->",special_script)
    # db.printing(debugging,arg_list)
    for arg in arg_list:
        
        if(arg != "<SWD>"):
            for ch in characters_to_be_removed:
                arg = arg.replace(ch, "")
            # db.printing(debugging,arg)
            if "0x" in arg:
                # db.printing(debugging,"found 0x")
                arg = arg.replace("0x", "")
                arg = int(arg, 16)
            elif 'L' in arg: #or 'H' in arg or 'X' in arg or 'C' in arg:  
                arg = convert_ate_string_to_int(arg)
            elif 'H' in arg or 'X' in arg or 'C' in arg: 
                arg = maskMAN(arg)
            else:
                #special_script flag will be set when the command such as <000000000A> is detected
                special_script = 1     
                # db.printing(debugging,"I am problem")
            cleaned_arg_list.append(arg)
    # db.printing(debugging,"inside->",special_script)
    return [cleaned_arg_list, special_script]

#This cleans up [WAIT] commands for delays during PFD parsing 
def clean_wait_args(pfd_line):
    wait_args = pfd_line
    characters_to_be_removed = "<>"
    for ch in characters_to_be_removed:
        wait_args = wait_args.replace(ch, "")
    wait_args = wait_args.split()
    delay_in_seconds = float(wait_args[0])
    number_of_elements = len(wait_args)
    if ( delay_in_seconds > 0.5 ) and (number_of_elements <= 1):    #---------------------test------------
        delay_in_seconds = 0.5
    #delay_in_seconds = .2 #have this here just for debugging purposes, need to comment out 
    number_of_elements = len(wait_args)
    if(number_of_elements >= 2):
        if(wait_args[1] == "ns"):
            delay_in_seconds = delay_in_seconds / NANOSECONDS_TO_SECONDS_CONVERSION_FACTOR
    if(number_of_elements >= 2):
        if(wait_args[1] == "ms"):
            delay_in_seconds = delay_in_seconds / MILLISECONDS_TO_SECONDS_CONVERSION_FACTOR
    return delay_in_seconds


def sleep(duration, get_now=perf_counter):
    now = get_now()
    end = now + duration
    # db.printing(debugging,now, end, duration)
    while now < end:
        now = get_now()

def set_vol_or_curr_for_shorted_pads(shorted_pads,ate_value_string):
    
    set_voltage_or_current(debugging,ate_value_string,shorted_pads[0])
    set_voltage_or_current(debugging,ate_value_string,shorted_pads[1])

    

    difference = DAQread(DAQ_PSYSP_PSYSN)
    
    counter = 0
    while (abs(difference) - 5e-4 >= 0):
        val = ate_value_string[:len(ate_value_string)-1]
        V_or_A = ate_value_string[-1]
        new_volt = float(val) + difference
        ate_value_string_new = str(new_volt)+V_or_A
        set_voltage_or_current(debugging,ate_value_string_new,shorted_pads[1])
        difference = DAQread(DAQ_PSYSP_PSYSN)
        db.printing(1,"**-> Applying: ",shorted_pads[0]," ",ate_value_string)
        db.printing(1,"**-> Applying: ",shorted_pads[1]," ",ate_value_string_new)
        # print(difference)
        counter+=1

        if counter>=5:
            report = db.get_global("running_report")
            temp = "Volt or current on Shorted pins is not correct"
            report['1'].append(temp)
            db.store_global("running_report",report)
            print('ERROR ERROR ERROR - UNCOMMENT in set_vol_or_curr_for_shorted_pads init')
            break
    db.printing(debugging,"**-> volt diff as read from DAQ = ",difference)
    
def send_ate_command(pfd_line,shorted_pads):
    global debugging,dict
    ate_args = pfd_line
    ate_args = ate_args.split() #splits strings as follows: ["Welcome", "To", "The", "Jungle"]
    cleaned_arg_list = []
    characters_to_be_removed = "[]<>:"
    for arg in ate_args:
        for ch in characters_to_be_removed:
            arg = arg.replace(ch, "")
        cleaned_arg_list.append(arg)
    ate_args = cleaned_arg_list
    #ate_args[1] to read second word
    if ("PAD" == ate_args[0]):
        if shorted_pads:
            
            ate_value_string = ate_args[2] 
            set_vol_or_curr_for_shorted_pads(shorted_pads,ate_value_string)


        else:
            ate_value_string = ate_args[2] 
            dict.update({ate_args[1]:ate_args[2][:(len(ate_args[2])-1)]})
            # print(dict)
            set_voltage_or_current(debugging,ate_value_string,ate_args[1])
            # db.printing(debugging,"@@@@@@",ate_args[1])



    #NOTE: Add code, if the code reads MEAS_V_PAD, create a new array and store varaiable
    #if the reading happens again, do not create a new array, store a new item and append
    elif "MEAS" in ate_args[0]:
        ate_range_string = ate_args[2]
        if "MEAS_V_PAD"  == ate_args[0]:
            #NOTE: PASS arugments to function : meas or force/channel etc 
            # print("atr_range: ",ate_range_string)
            # print("ate_args: ",ate_args)
            measure_voltage_and_current(ate_range_string,ate_args)
            # read_volt_from_smu('1')
            # print(dict)
        elif "MEAS_F_PAD" == ate_args[0]:
            db.printing(debugging,"Meas_F_Pad Detected")
            #measure_frequency(ate_range_string)
    else:
        db.printing(debugging,"Input PAD string is invalid")


#To enter test mode during PFD parsing, type the command [TEST] within the PFD
def make_a_breakpoint(line):
    print("*"*20,"Waiting at breakpt","*"*20)
    testing = 1
    # bar.pause()
    end_this_run = 0
    while testing:
        try:
            msg = """ Enter key to cont.
            \t . - end this pfd 
            \t 0 - continue as is 
            \t 1 - Read a register 
            \t 2 - Write a register
            \t 3 - Read All Trim Reg
            \t 4 - Save trim reg to xl (after '3')
            \t 5 - read all trims from this pfd
            \t 6 - measure voltage of DAQ (eg. 210)
            \t 7 - save to excel
            =>"""
            breakpt = input(msg)
            print(breakpt)
            if breakpt == '0':
                testing = 0
                end_this_run = 2
                print("=> Continuing->")
            elif breakpt == '.':
                try:
                    save_in_excel()
                except:
                    pass
                end_this_run = 1
                break
            elif breakpt =='1':
                reg_name = input("Enter reg address (eg- trim.BG_TRIM2.sp_reg_lo_trim, 0x70000000)\n\n=>")
                print("- - - - - - - -")
                print("**-> Reading Address: \t\t", reg_name)
                try:
                    if '0x' in reg_name:
                        adressNum = int(reg_name[2:],16) #interger value of the adress here 
                        read_data = db.jlink.memory_read32(adressNum,1)
                        print("**-> Register reading is: \t" , bin(read_data[0])[2:]," <=> ",hex(read_data[0]))
                        print("- - - - - - - -")
                    else:
                        read_reg_data(db.jlink,1,reg_name)
                except:
                    print("Check input again")
            elif breakpt =='2':
                reg_name = input("Enter reg name (eg- trim.BG_TRIM2.sp_reg_lo_trim(write masked),0x70000000 (full reg write))\n=>")
                reg_data = input("Enter reg data (eg- 5, 0x30106b9d | data masked if full reg name entered)\n=>")
                if ("0x" in reg_data):
                    reg_data = int(reg_data[2:],16)
                    print(reg_data)
                if ("0x" in  reg_name):
                    reg_name = int(reg_name[2:],16)
                    print("reg_name=> ",reg_name)
                    print("reg_data=>",reg_data)
                    db.jlink.memory_write32(reg_name, [int(reg_data)])
                    print("written, now reading")
                    read_value = db.jlink.memory_read32(reg_name,1)
                    if (int(read_value[0]) != int(reg_data) ):
                        print("Error writing register\nWritten value=",reg_data,"\nRead value after write=",read_value[0])

                    # print(read_value[0])
                else:
                    findREG(db.jlink,1,reg_name, int(reg_data))
                print("- - - - - - - -")
            elif breakpt == '3':
                read_and_save_all_trims(0)
            elif breakpt == '4':
                save_new_trims_to_xl(line.replace("[TEST]",''))
            elif breakpt == '5':
                pfd_name = db.get_global("pfd_name")
                read_all_trim_reg_from_pfd(pfd_name)
            elif breakpt == '6':
                pad_name = input("Enter pad_name")
                print(pad_name)
                if pad_name.isnumeric():
                    read = DAQread(pad_name)
                    print(read,"V")
            elif breakpt == '7':
                save_in_excel()
        except:
            print("=> try again")
    return end_this_run


#VGPTAT reading works with DAQ, but not with smu, maybe channel 1 is getting turned off during the read 
def measure_voltage_and_current(test,ate_args):
    global elez, debugging
    res = ate_args[4] #save result to
    # print(res)
    db.printing(debugging,"YOUR VARIABLE READING IS:", res)
    CH1_DAQ = db.get_global("CH1_DAQ")
    CH2_DAQ = db.get_global("CH2_DAQ")
    CH3_DAQ = db.get_global("CH3_DAQ")
    pads = {"pad_avrrdy" : CH1_DAQ, "pad_avren" : CH2_DAQ, "vd12" : CH3_DAQ}
    check_avren_list = ["VBE_100uA_p","VBE_100uA_n","VBE_5uA_n","VBE_5uA_p","V_1K_p","V_1K_n"]
    if VGPTAT_FLAG == 1:
        if res in check_avren_list:
            elez = DAQread(CH2_DAQ) #1p11 5,6
            db.printing(debugging, "reading pad=>","pad_avren")
        else:
            elez = DAQread(CH1_DAQ)
            db.printing(debugging, "reading pad=>","pad_avrrdy")
        #res = ate_args[4]
        dict.update({res: elez}) #Fiilling up dictionary here To save into Excel
        
        try:
            global Target_VPtat_temp
            delta_vbe =  (dict["VBE_100uA_p"] - dict["VBE_100uA_n"]) - (dict["VBE_5uA_p"] - dict["VBE_5uA_n"])
            V_1K =  dict["V_1K_p"] - dict["V_1K_n"]
            Res_1K = V_1K/100e-6
            Target_VPtat_temp = 0.002679*((delta_vbe-0.072253)/0.000261)+0.579
            db.printing(debugging,"**-> delta_vbe temp is: ",delta_vbe," <= (",dict["VBE_100uA_p"],' - ', dict["VBE_100uA_n"],') - (',dict["VBE_5uA_p"],' - ',dict["VBE_5uA_n"])
            db.printing(debugging,"**-> V_1K: ",V_1K," <= ",dict["V_1K_p"],' - ',dict["V_1K_n"])
            db.printing(debugging,"**-> Res_1K: ",Res_1K," <= ",V_1K,"/100e-6")
            db.printing(1,"**-> VPTAT_Target : " , Target_VPtat_temp)

            db.printing(debugging,"**-> Res_1k : ", Res_1K)
 
        except:
            db.printing(debugging,"**-> All [VBE_100/5uA, V_1K] values not found yet, continue in pfd, Target_VPtat not defined yet") 
    
    elif VDDLO_FLAG ==1:
        elez = DAQread(CH1_DAQ)
        VDDLO_trim.append(data2)
        VDDLO_BANK.append(elez)
    elif VDDA_FLAG ==1:
        elez = DAQread(CH1_DAQ)
        VDDA_trim.append(data2)
        VDDA_BANK.append(elez)
    elif IS_LR_FLAG ==1:
        elez = DAQread(CH1_DAQ)
        IS_LR_trim.append(data2)
        IS_LR_BANK.append(elez)
        #Place Second DAQ Channel on VD12
    elif VDDLR_FLAG ==1:
        elez = DAQread(CH3_DAQ) #1p9 3,4
        VDDLR_trim.append(data2)
        VDDLR_BANK.append(elez)
    elif VDDSCR_FLAG ==1:
        elez = DAQread(CH3_DAQ)
        VDDSCR_trim.append(data2)
        VDDSCR_BANK.append(elez)



    elif res =="tmux_sp_vdd_scr_code_final" or res == "tmux_sp_vdd_lr_code_final":
        db.printing(debugging,"LOOKING AT VD12")
        elez = DAQread(CH3_DAQ)
    elif res.isnumeric():
        print(res)
        elez = DAQread(res)
    else: 
        db.printing(debugging,"Normal TMUXP Read")
        try:
            temp = ate_args[1].split(',')
            daq1 = float(DAQread(pads[temp[0]]))
            daq2 = float(DAQread(pads[temp[1]]))
            elez = daq1-daq2
            db.printing(debugging,"**-> ",daq1," - ",daq2)
        except:
            if res == "vd12":
                elez = DAQread(CH3_DAQ)
            else:
                elez = DAQread(CH1_DAQ)
            # elez = DAQread(CH3_DAQ)
    db.printing(debugging,"YOUR DAQ READING IS:", elez)
    db.printing(1,"-> MEAS_V_PAD->",res,"-"*(50-len(res))," -> ",elez)
    dict.update({res: elez}) #Fiilling up dictionary here To save into Excel
    return [res,elez]


#Frequency not measured in any PFDS at the moment, this function has no purpose 
def measure_frequency(ate_range_string):
    global debugging
    db.printing(debugging,ate_range_string)
    ate_range_string = ate_range_string.replace("Hz", "")
    ate_range = ate_range_string.split(',', 1)
    min_freq = float(ate_range[0])
    max_freq = float(ate_range[1])
    db.printing(debugging,min_freq, max_freq)
  

#Search the DATA BASE for any variable name similar, if nothing matches, perform a special script 
def write_swd(jlin,swd_args):
    global VDDLO_trimAdress,VDDA_trimAdress,VDDSCR_trimAdress, VDDLR_trimAdress,IS_LR_trimAdress
    global data2, debugging
    db.printing(debugging,"**-> Write: Addr: %s, Data: %s" %(hex(swd_args[0]), hex(swd_args[1]))) 
    if jlink_ENABLE and special_script ==0:
        data = [swd_args[1]]
        db.jlink.memory_write32(swd_args[0], data)
    #trim values are manually done from register writes in the PFD< so code below is used for trimming purposes 
    if IS_LR_FLAG ==1:
        IS_LR_trimAdress = swd_args[0]
        data2 = data[0] 

    if VDDLO_FLAG ==1:
        VDDLO_trimAdress = swd_args[0]
        data2 = data[0]   
        # db.printing(debugging,"**->WR: Addr: %s, Data: %s" %(hex(swd_args[0]), hex(swd_args[1])))

    if VDDA_FLAG ==1:
       VDDA_trimAdress = swd_args[0]
       data2 = data[0]   
    # db.printing(debugging,"WR: Addr: %s, Data: %s" %(hex(swd_args[0]), hex(swd_args[1]))) 

    if VDDSCR_FLAG ==1:
       VDDSCR_trimAdress = swd_args[0]
       data2 = data[0]   
    # db.printing(debugging,"WR: Addr: %s, Data: %s" %(hex(swd_args[0]), hex(swd_args[1])))

    if VDDLR_FLAG ==1:
       VDDLR_trimAdress = swd_args[0]
       data2 = data[0]    
    #move code below to the MEAS_V_PAD section 


def read_swd(swd_args,regcapture,saveMask): #only need to read here, dont need to store any variables 
    if jlink_ENABLE:
        #hi = masking bits 
        global maskBits, dict
        global snew2, debugging
        saveMEPLZ = swd_args[0]
        read_data = db.jlink.memory_read(swd_args[0], 4) #reads 4 bytes of the register adress
        read_data = read_data[::-1] #reversing using list slicing
        read_data[0]= '{0:08b}'.format(read_data[0]) #list elements needs to be in 8 bit format or the MASK will not work 
        read_data[1]= '{0:08b}'.format(read_data[1])
        read_data[2]= '{0:08b}'.format(read_data[2])
        read_data[3]= '{0:08b}'.format(read_data[3])
        s = ' '.join([str(n) for n in read_data]) #combining all ones and zeros together here
        
        # db.printing(debugging,"The register reading is:", s)
        db.printing(debugging,"**-> Reading: Add: ",hex(swd_args[0])," Data: ",s)
        s = s.replace(" ", "")
        db.printing(debugging,"**-> Data:\t",s)
        b = 0 
        print(saveMask)

        if len(regcapture) != 0: #checks to see if there is anything inside the reg_capture array
          #for i in saveMask:
                first_value = list(saveMask.values())[b] #will iterate through the saveMask list 
                first_value= first_value.replace(" ", "")  # This will get rid of the spaces in the register name
                # db.printing(debugging,"The mask you want to use is:",first_value)
                num_leading = len(first_value) - len(first_value.lstrip('0'))
                num_trailing = len(first_value) - len(first_value.rstrip('0'))
                num_trailing = num_trailing*-1
                # db.printing(debugging,"leading characters to remove is:" , num_leading)
                # db.printing(debugging,"trAILING characters to remove is:" , num_trailing)
                snew = s[num_leading:]  # remove the first two character
                # db.printing(debugging,"Removing the leading zeros gives: ", snew)



                if num_trailing==0:
                    snew2 = snew  # remove the first two character
                else:
                    snew2 = snew[:num_trailing]
                # db.printing(debugging,snew2)
                # db.printing(debugging,"Removing the trailing zeros gives: ", snew2)

                snew2 = int(snew2,2)
                ########################################################################
                #dataRegisterName.append(regcapture)
                print("The data you are reading into excel is -->",snew2)
 
                dataReadArray.append(snew2)


                array = [dataRegisterName,dataReadArray]
                df = pd.DataFrame(array).T
                df.to_excel(excel_writer = "C:/Users/TorreroSanch/Desktop/ReadRegistersPFD.xlsx")

                ########################################################################
                # db.printing(debugging,snew2)
                # db.printing(debugging,"Number of characters removed in front", num_leading)

                newSave.update({regcapture[b]: snew2}) #this is where well save all the register read data 
                this_var = regcapture[b].split('.')[1]
                dict.update({this_var: snew2})
                #code above need sto be changed to store it in the OG dict

               # b=b+1
            # db.printing(debugging,"Your final testing dictionary is now:", newSave)
            ###############################################################3
        # db.printing(debugging,dict)
        regcapture =[] #reseting the registers in case of other reg_capture messages
        saveMask = {} #reseting the saveMask, final mask is called newSace


def read_swd_simple(swd_args): #only need to read here, dont need to store any variables 
    if jlink_ENABLE:
        #hi = masking bits 
        global maskBits
        global snew2, debugging
        saveMEPLZ = swd_args[0]
        db.printing(debugging,"The Register you want to read from is --->: 0x" , hex(swd_args[0]))
        read_data = db.jlink.memory_read(swd_args[0], 4) #reads 4 bytes of the register adress
        read_data = read_data[::-1] #reversing using list slicing
        read_data[0]= '{0:08b}'.format(read_data[0]) #list elements needs to be in 8 bit format or the MASK will not work 
        read_data[1]= '{0:08b}'.format(read_data[1])
        read_data[2]= '{0:08b}'.format(read_data[2])
        read_data[3]= '{0:08b}'.format(read_data[3])
        s = ''.join([str(n) for n in read_data]) #combining all ones and zeros together here
        s = s.replace(" ", "")
        db.printing(debugging,"$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$")
        db.printing(debugging,"$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$")
        db.printing(debugging,"The register reading is:", s)
        db.printing(debugging,"$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$")
        db.printing(debugging,"$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$")


from read_register import check_if_untrim
def write_trim_to_otp(otp_file_name):
    print("\n","-"*20,"Reading trims before power cycle","-"*20)
    
    read_and_save_all_trims(debugging)  # this line is just to save in python and display on terminal

    print("\n","-"*20,"Writing trims to OTP","-"*20)
    
    parse_pfd_file(otp_file_name)


    sleep(0.2)

    read_and_save_all_trims(debugging)  # save once here and then compare after POW
    save_new_trims_to_xl("","Before power cycle",0) 

    db.jlink.close()

    full_power_cycle() # turn off firmware also
    # power_cycle() # 

    # parse_pfd_file("ate_fw_stop_kernel.pfd")
    # parse_pfd_file("download_trim.pfd")

    db.jlink.open()
    db.jlink.set_tif(pylink.enums.JLinkInterfaces.SWD)
    db.jlink.connect('CORTEX-M0')


    print("\n","-"*20,"Check Trims after power cycle","-"*20)

    are_trims_correct = compare_if_trims_correct_after_otp(debugging)

    if (are_trims_correct == 0):
        db.printing(1,"-OTP-ERROR-"*15)

        report = db.get_global("running_report")
        temp = "Error in Writing to OTP, trim registers before and after power cycle do not match"
        report['1'].append(temp)
        db.store_global("running_report",report)
        db.report_me_i_did_something(debug_option['email'])
        make_a_breakpoint('')


    # for i in range(0,2):
    #     # read_register_data(debugging,address,reg_bits,db.jlink) 
    #     check_if_untrim(db.jlink,0)  
    #     # Read_BG_volt(1)
    # sleep(2)

    

def update_globals(debug,debug_options,target,pfd_folder_location): # globals for calculations and printing
    global debugging,debug_option,target_global
    global BG_Target,VDDLO_Target,VDDLR_Target,VDDA_Target,VDDSCR_TARGET,TS_current_target,IS_LR_Target
    debug_option.update(debug_options)
    for i in target:
        target_global.update({i:target[i]})
    debugging = debug

    db.store_global("register_map",debug_options["register_map"])
    db.store_global("check_for_subsection",debug_options["check_for_subsection"])
    db.store_global("predefined_section",debug_options["predefined_section"])
    db.store_global("target_global",target_global)
    BG_Target=target_global["BG_Target"]
    VDDLO_Target=target_global["VDDLO_Target"] 
    VDDLR_Target=target_global["VDDLR_Target"]
    VDDA_Target=target_global["VDDA_Target"]
    VDDSCR_TARGET=target_global["VDDSCR_TARGET"] 
    IS_LR_Target = target_global["IS_LR_target"]
    TS_current_target=target_global["TS_current_target"] 
    db.store_global("pfd_name",'parsing_pfd')
    db.store_global("pfd_folder",pfd_folder_location)
    
    db.store_global("email",debug_option["email"])
    
import shutil
def do_initial_setup(pfd_folder_location,debug_options):

    db.refresh_this_log('parsing_pfd')
    # db.store_global("running_report",{'0':[],'1':[]})
    db.store_global("register_map",debug_options["register_map"])

    T_now = datetime.now()
    date = str(T_now.strftime("%m-%d-%y"))
    # print(date)

    directory_path = os.getcwd()
    temp = directory_path+"\\"+pfd_folder_location
    os.chdir(temp)
    directory_path = os.getcwd()

    temp3 = temp+"\\"+"All_results"
    if not os.path.exists(temp3):   # create folder for storing old excel files
        os.mkdir("All_results")


    temp1 = temp+"\\"+"results.xlsx"
    if not os.path.isfile(temp1):   # create results.xlsx if not exist
        wb = Workbook()
        print("Creating new Excel file results.xlsx for storing results")
        wb.save("results.xlsx")

    temp4 = temp+"\\All_results\\"+date+"-results.xlsx"
    if not os.path.exists(temp4):   # store excel file
        print("Creating backup of results.xlsx, storing in All_results")
        shutil.copy(temp+"\\"+"results.xlsx",temp4)

    temp2 = temp+"\\"+"run_log"
    if not os.path.exists(temp2):   # create folder for storing log files
        os.mkdir("run_log")
        
    temp = ("../"*len(pfd_folder_location.split("\\")))
    os.chdir(temp)
    save_in_excel() # will create appropriate sheets if we made a new file 


def read_pfd_files(PFD_FOLDER,debug,debug_options,target):
    print("\n")
    update_globals(debug,debug_options,target,PFD_FOLDER)
    do_initial_setup(PFD_FOLDER,debug_options)
    directory_path = os.getcwd()
    global pfd_folder,pfd_name,all_pfd
    pfd_folder = PFD_FOLDER
    os.chdir(directory_path+"\\"+PFD_FOLDER)
    directory_path = os.getcwd()

    # db.sharing_globals('parsing_pfd',PFD_FOLDER)

    # db.refresh_this_log('parsing_pfd')

    for file in os.listdir():
        # print(file)
        if file.endswith(".pfd"):
            global pfd_num    
            file_PATH = str(file)
            pfd_name.append(file)
            f =  open(file_PATH, 'r')
            pfd_file =f.read() 
            this_pfd = pfd_file.split('\n')
            # while("" in this_pfd) :
            #     this_pfd.remove("")
            all_pfd.append(this_pfd)
            # pfd_num +=1
            f.close()
        else:
            # db.printing(debugging,"***not a pfd file =>",file)
            pass
            # db.printing(debugging,file)
    os.chdir("../"*len(pfd_folder.split("\\")))
    print("\n")

    return [all_pfd,pfd_name]

dict_of_subsections = {}
def find_all_subsections_in_pfd(this_pfd,look_for_subsections,is_section_name_ok):
    global dict_of_subsections
    keep_adding=-1
    all_section_name = []
    this_section_lines = []
    what_to_run = ''
    for i in this_pfd:
        this_instr = i.split(' ')[0]
        # print(this_instr)
        # print(i)
        if ('SUBSECTION-' in i) and (this_instr == "[MSG]") :
            
            if ("START" in i):
                keep_adding = 1
                this_section_name = i.split(' ')[-1]
                all_section_name.append( this_section_name )
            elif("END in i"):
                keep_adding = 0
        if (keep_adding == 1):
            this_section_lines.append(i)
        elif (keep_adding == 0) and ( this_section_name not in dict_of_subsections):
            this_section_lines.append(i)
            dict_of_subsections.update({this_section_name:this_section_lines})
            this_section_lines = []
    if (look_for_subsections==1) and (is_section_name_ok==0): # first try from debug_options if wrong then come here
        db.printing(1,"**-> Which subsections to run ( eg.=>  sub1 sub3 sub2  (run in order))")
        # print(dict_of_subsections)
        for i in dict_of_subsections:
            db.printing(1,"\t-",i)
        what_to_run = input('\n')

    what_to_run = what_to_run.split(' ')
    return [what_to_run,all_section_name]
    


is_var = {}
def parse_and_store_equation(line,calcdict,jlink):
    global debugging,is_var, dict
    # print(dict)
    pfd_name = db.get_global("pfd_name")
    db.printing("-"*100)
    this_line = re.split(":|;",line)
    equation = this_line[1:]
    format_check= ''

    store_as_int = 0
    if ('(uint32_t)' in line.split('=')[1]):
        store_as_int = 1

    equation = [s.replace('(uint32_t)','').replace('(double)','') for s in equation]

    this_eqns = {}
    this_result = {}
    for i in range(len(equation)-1):
        this_eqns[i] ={equation[i].split("=")[0].replace(" ",""):equation[i].split("=")[1].replace(" ","")} 

    for j in range(1,len(equation)-1):
        i=j # testing other loop
        eqn_var = (f"{(equation[i].split('='))[0]}").replace(" ",'')
        read_reg = (equation[i].split('='))[1]
        db.printing(debugging,"------ Next sub equation ------")
        db.printing(debugging,"**-> Reading: REG: ",read_reg)
        try:
            # print(is_var)
            # print(calcdict)
            # print("trying read reg")
            if ('uint32_t' in line):    # and ('accuracy' not in pfd_name):
                read_value = read_reg_data(jlink,debugging,read_reg)
            elif ('is' in pfd_name) and ('accuracy' in pfd_name):
                val = is_var[read_reg]
                if (val[0]=='-'):
                    read_value = -1*int(is_var[read_reg][3:],16) 
                else:
                    read_value = int(is_var[read_reg][2:],16)
            # print(read_value)
            [reg_format, decim, fracti] = read_reg_sign_bits(jlink,debugging,read_reg)
            # print("format:",reg_format)
            # print()
            db.printing(debugging,"**-> Data: ",bin(read_value)," <=> ",hex(read_value)," <=> ", int(read_value))
            ############### FOR READING -ive NUMBER ########################
            # if( int(read_value)>=128) and ("0x" not in equation) :
            #     if ("accuracy" not in pfd_name):
            #         read_value=Bits(bin = bin(read_value)[2:]).int
            #     else:
            #     # print(read_value)
            #       if( len(bin(read_value)[2:]) == mask_width  ) and ("0x" not in equation):
            #             read_value=Bits(bin = bin(read_value)[2:]).int
            # print("format: ",reg_format, decim, fracti)
            # print(equation)
            # print(reg_format)
            # print(decim)
            # print(read_value)
            # print(bin(read_value))
            # print(bin(read_value)," ",len(bin(read_value)[2:])," ",decim+fracti)
            if (reg_format == "exception") and ( int(read_value)>=128):
                read_value=Bits(bin = bin(read_value)[2:]).int
                format_check = '(format)'
                
            elif (len(bin(read_value)[2:]) == decim+fracti ):# and ('0x' not in equation):
                read_value=Bits(bin = bin(read_value)[2:]).int
            
            
        except:
            db.printing(debugging,"**-> Register name not found")
                
        # exception_dict.update({"pad_isen1":2,"v_one_k_1":1})#################---test remove from here
            try:
                try:
                    db.printing(debugging,"**-> Trying to resolve --> Looking in variables instead of register name")
                    read_value = exception_dict[read_reg]
                    db.printing(debugging,"**-> *DATA FOUND*, using it")
                    # try:
                    #     db.printing(debugging,"**-> Data: ",bin(read_value)," <=> ",hex(read_value)," <=> ", int(read_value))
                    # except:
                    #     db.printing(debugging,"**-> Data: ",read_value)
                except:
                    if ("is" in this_file_name ) or ("test" in this_file_name):
                        db.printing(debugging,"**-> Looking in IS variables")
                        try:
                            db.printing("not in is var")
                            read_value = is_var[read_reg]
                        except:
                            try:
                                db.printing("not in dict")
                                read_value = dict[read_reg]
                                
                            except:
                                if ("pad_isen" in read_reg) and ("accuracy" in this_file_name):
                                    read_value = float((read_reg.split("_")[-1]).replace("uA",""))*1e-6
                                    if (read_value==0):
                                        read_value = 0.001*1e-6
                                    db.printing(debugging,"read_isen= ",read_value)
                                else:
                                    db.printing(1,"*"*100,"Error data not found, test_prabal_init-> parse_and_store_equation")
                               
            except:
                db.printing(debugging,"**-> Error, data not found")
        try:
            db.printing(debugging,"**-> Data: ",bin(read_value)," <=> ",hex(read_value)," <=> ", int(read_value))
        except:
            db.printing(debugging,"**-> ",read_reg," : ",read_value)
        this_result.update({eqn_var:read_value})
        # 
    for var in this_result:
        final_var = ((equation[0].split('='))[0]).replace(" ",'')
        try:
            calc = (((equation[0].split('='))[1]).replace(" ",'')).replace(var,str(this_result[var])).replace("uV","").replace("mV","")
        except:
            print ("-ERROR-"*20)
        
        #
        # try:
        #      calc = (((equation[0].split('='))[1]).replace(" ",'')).replace(var,str(int(this_result[var]))).replace("uV","").replace("mV","")
        # except:
        #     print ("-ERROR-"*20)

        equation[0] = final_var+"="+calc
        
        
  
    # 
    # db.printing(debugging,calc)
    if '0x' in calc:
        calc = "hex("+calc.replace("*"," & ")+")"
    ### check for mV,uV, format of calc before storing
    ### convert to unit32_t for all using numpy
    # db.printing(1,"**-> Solving equation=>",equation)
    db.printing(debugging,"**-> Solving equation=>",calc)
    # print(calc)
    try:
        calc = eval(calc)
    except:
        db.printing(debugging,"**-> maybe parenthesis error, trying parenthesis")    
    
        try:
            # db.printing(debugging,"**-> Solving equation=>",calc)
            calc = eval(calc+")")
            
        except:
            
            db.printing(1,"************************** Equation parsing error in line or divide by 0  **********************")
            db.printing(1,"************************** Taking 1e-10 as result  **********************")
            calc = 1e-10
            db.printing(debugging,line)
    if calc == 0.0: # no floating point 0
        calc = 0        
    if ("is" in this_file_name ) or ("test" in this_file_name):
        # if ('accuracy' in this_file_name):
        #     is_acc_var.update({final_var:calc})
        is_var.update({final_var:calc})  
        # 

    # db.printing(debugging,"**->", final_var," = ",calc)  
    # if type(calc) is not float:
    if store_as_int==1:# ("(uint32_t)" in line and "(double)" not in line) and ('0x' not in str(calc)): #
        if (type(calc) == str) and ('0x' not in calc):
            calc = round(float(calc))
        elif type(calc) == float:
            calc = round(calc)
        elif (type(calc) == str) and ('0x' in calc):
            calc = round(int(calc,16))
        
    this_pfd = db.get_global("pfd_name")
    
    if ("vs_adc" in this_pfd): 
        # this_val = int(bin(calc + 2**13)[2:],2)
        this_val = calc
        db.printing(1,"-> ",final_var,"-"*(50-len(final_var))," -> ",(this_val/8)*0.00125) # for conversion of read vaue to voltage
        # print(this_val)
        # (this_val/8)*0.00125
    else:
        if ('.' in final_var):
            # db.printing(1,"-> ",final_var,"-"*(50-len(final_var))," -> ",hex(calc)," ",format_check)
            db.printing(1,"-> ",final_var,"-"*(50-len(final_var))," -> ",hex(calc)," ",format_check)
            # print("-> ",final_var,"-"*(50-len(final_var))," -> ",hex(calc))
        else:
            db.printing(1,"-> ",final_var,"-"*(50-len(final_var))," -> ",calc)
            # print("-> ",final_var,"-"*(50-len(final_var))," -> ",calc)
    # print ("**-> ",final_var,"\t:\t",calc)
    calcdict.update({final_var:calc} )    


def find_and_save_computed_value(line,calcdict,jlink):
    global write_calculated_flag, exception_dict, debugging


    if ("Restore the values back to the general purpose registers" in line):
        return [int('0',16),write_calculated_flag] # reset to 0, IS.pfd at the end

    eqn = re.split(":|=",line)
    register = eqn[1].replace(" ","")
    eqn = eqn[2].replace("(","( ").replace("<"," < ").replace(register+".","").replace(")"," )")
    each_word = eqn.split(" ")
    # db.printing(debugging,each_word)
    # fix_for_ts = {"tsen_adc_src4_ate_gain":}
    for i in range(len(each_word)):
        # db.printing(debugging,each_word[i])
        if each_word[i].find('x')==-1 and each_word[i]!='':
            # db.printing(debugging,each_word[i])
            if(each_word[i][0] == '0'):
                each_word[i] = each_word[i][1:]
        
        if each_word[i] in calcdict:
            # db.printing(debugging,calcdict[each_word[i]])
            # db.printing(debugging,calcdict)
            each_word[i] = str(hex(int(calcdict[each_word[i]])))
            # db.printing(debugging,each_word[i])

        if "vs0_os_trim_bif_ana" in calcdict:
            exception_dict.update(vs0_os_trim_bif_ana = calcdict["vs0_os_trim_bif_ana"])
            calcdict.update({"vs0_os_trim_bif":calcdict["vs0_os_trim_bif_ana"]})
        if "vs1_os_trim_bif_ana" in calcdict:
            exception_dict.update(vs1_os_trim_bif_ana = calcdict["vs1_os_trim_bif_ana"])
            calcdict.update({"vs1_os_trim_bif":calcdict["vs1_os_trim_bif_ana"]})

    # db.printing(debugging,eqn)
    # db.printing(debugging,each_word)
    eqn = "".join(each_word)
    db.printing(debugging,"**-> ",register,"=",eqn)
    eqn = eval("hex("+eqn.replace(" ","")+")")
    # db.printing(debugging,eqn)
    # db.printing(debugging,int(eqn[2:],16))
    db.printing(debugging,"**-> ",register,"=",eqn," <=> ",int(eqn[2:],16))
    write_calculated_flag = 1
    # db.printing(debugging,"wc_flag",write_calculated_flag)
    print(calcdict)
 
    df = pd.DataFrame(data=calcdict, index=[0])

    df = (df.T)
    df.to_excel('C:/Users/TorreroSanch/Desktop/ReadTrimDictionary.xlsx')

    return [int(eqn[2:],16),write_calculated_flag]
    # write_full_reg("trim.VSP_GAIN_TRIM0.vsp0_adc_gain_vid2", eqn,db.jlink)
    # db.printing(debugging,register)



def read_all_trim_reg_from_pfd(this_pfd_name):
    # print ("Resetting")
    global pfd_name
    number = pfd_name.index(this_pfd_name)
    db.store_global("pfd_name",this_pfd_name)
    db.store_global("pfd_folder",pfd_folder)
    all_trim_reg_dict = db.get_global("all_trim_regs")
    trim_values = {}
    line_no = 1
    seperate_reg = []
    write_flag = 0
    write_add_data = [0,0]
    db.printing(1,"**-> Reading trims from ", this_pfd_name)
    with alive_bar(len(all_pfd[number]),title="Reading trim reg") as bar:

        for line in all_pfd[number]:
            bar()
            pfd_line = line.split(" ",1)
            if "" == line:
                # db.printing(debugging,"**-> skipping line for reset register")
                pass
            elif ('reg_write_computed' in line) and ('=' in line) and ('|' in line):
                reg_size = line.split('=')[1].split("|")[1].split("<<")[1].split(")")[0]
                reg_size = int(reg_size)

                reg_name = line.split('=')[0].split(':')[1].split('.')[-1]
                # print(reg_name)
            elif "[WRDP]" == pfd_line[0] or "[WRAP]" == pfd_line[0]:
                
                write_add_data[write_flag] = line.split(" ")[-1].replace("<","").replace(">","") 
                write_flag +=1
                if write_flag == 2:
                    write_flag =0
                    if (len(write_add_data[1]) > 10) and ("0xe" not in write_add_data[0] ):
                        
                        reg_add = int(write_add_data[0].replace('0x',''),16)
                        # read_value = hex(db.jlink.memory_read32(reg_add,1)[0])

                        read_data = db.jlink.memory_read(reg_add, 4) #reads 4 bytes of the register adress
                        read_data = read_data[::-1] #reversing the list using list slicing
                        #read_data = ["{0:b}".format(i) for i in read_data]
                        read_data[0]= '{0:08b}'.format(read_data[0]) #list needs to be in 8 bit format or the MASK will not work 
                        read_data[1]= '{0:08b}'.format(read_data[1])
                        read_data[2]= '{0:08b}'.format(read_data[2])
                        read_data[3]= '{0:08b}'.format(read_data[3])

                        # print(seperate_reg)
                        temp = ''
                        for i in seperate_reg:
                            temp = temp+' ' +str(int(i,2))

                        read_value = ' '.join(read_data)
                        read_value_b = format(db.jlink.memory_read32(reg_add,1)[0], '032b')
                        count = 0
                        this_sub_reg = ''
                        for i in range(len(read_value_b)):
                            j = len(read_value_b)-i-1
                            count +=1
                            this_sub_reg = read_value_b[j] + this_sub_reg
                            # print(read_value_b[j])
                            if count == reg_size:
                               seperate_reg.append(this_sub_reg)
                               this_sub_reg = '' 
                               count =0
                            
                        # print(read_value_b)
                        seperate_reg.reverse()
                        for i in range(len(seperate_reg)):
                            if len(seperate_reg[i])>=4:
                                seperate_reg[i] =str(Bits(bin = seperate_reg[i]).int)
                            else:
                                seperate_reg[i] = str(Bits(bin = '0'+seperate_reg[i]).int)
                        temp = ' '.join(seperate_reg)
                        read_value = read_value+" ( "+temp+" )"
                        write_add_data[0] = write_add_data[0] + " ("+str(reg_size) +" bits)"
                        db.printing(1,"**-> \t",write_add_data[0],"=> \t",read_value)

                        # db.printing(1,"**-> \t",write_add_data[0],"=> \t",read_value,"\t<=>\t",bin(int(read_value[2:],16)))
                        # read_value = hex(db.jlink.memory_read32(reg_add,1)[0])
                        all_trim_reg_dict.update({write_add_data[0]:[0,read_value,reg_name]})
                        temp = " ( "+temp+" )"
                        all_trim_reg_dict.update({write_add_data[0]:[0,temp,reg_name]})
                        seperate_reg = []
                        # db.jlink.memory_write32( reg_add, [0]   )
                    write_add_data = [0,0]

    db.store_global('all_trim_regs',all_trim_reg_dict)


def compare_if_trims_correct_after_otp(debugging):
    ## use read_and_save_all_trims before power_cycle, then use this to compare wwith those values
    are_trims_correct = 1
    old_trimmed_dict = db.get_global('all_trimmmed_dict')

    read_and_save_all_trims(debugging)  # 
    save_new_trims_to_xl("","After power cycle",1)

    new_trimmed_dict = db.get_global('all_trimmmed_dict')

    error_dict = {}

    for i in old_trimmed_dict:
        if i in new_trimmed_dict:
            db.printing(debugging,"**->  ",i,'-'*(60-len(i)), old_trimmed_dict[i][1], '(before PWR)', '---->', new_trimmed_dict[i][1], "(After PWR)" )
            
            if (old_trimmed_dict[i][1] != new_trimmed_dict[i][1]):
                error_dict.update({i : [old_trimmed_dict[i], new_trimmed_dict[i] ]})
                print(error_dict)

        else:
            db.printing(1,"*-> ", i , " got reset") 
            are_trims_correct = 0

    
    if error_dict:
        db.printing(1,"\n\n**-> Trims got changed")
        are_trims_correct = 0
        for i in error_dict:
            db.printing(1,"**->  ",i,'-'*(60-len(i)), old_trimmed_dict[i][1], '(before PWR)', '---->', new_trimmed_dict[i][1], "(After PWR)" )
    # else:
    #     db.printing(1,"\n*****-> Trims are same as before PWR <-*****")
    return are_trims_correct




def reset_all_trim_reg_from_pfd(this_pfd_name):
    # print ("Resetting")
    global pfd_name
    number = pfd_name.index(this_pfd_name)
    db.store_global("pfd_name",this_pfd_name)
    db.store_global("pfd_folder",pfd_folder)
    # all_trim_reg_dict = db.get_global("all_trim_regs")
    trim_values = {}
    line_no = 1
    write_flag = 0
    write_add_data = [0,0]
    db.printing(1,"**-> Resetting trims from ", this_pfd_name)
    with alive_bar(len(all_pfd[number]),title="Resetting trim reg") as bar:

        for line in all_pfd[number]:
            bar()
            pfd_line = line.split(" ",1)
            if "" == line:
                # db.printing(debugging,"**-> skipping line for reset register")
                pass
            elif "[WRDP]" == pfd_line[0] or "[WRAP]" == pfd_line[0]:
                
                write_add_data[write_flag] = line.split(" ")[-1].replace("<","").replace(">","") 
                write_flag +=1
                if write_flag == 2:
                    write_flag =0
                    if len(write_add_data[1]) > 10:
                        
                        # reg_add = int(write_add_data[0].replace('0x',''),16)
                        # # read_value = hex(db.jlink.memory_read32(reg_add,1)[0])

                        # read_data = db.jlink.memory_read(reg_add, 4) #reads 4 bytes of the register adress
                        # read_data = read_data[::-1] #reversing the list using list slicing
                        # #read_data = ["{0:b}".format(i) for i in read_data]
                        # read_data[0]= '{0:08b}'.format(read_data[0]) #list needs to be in 8 bit format or the MASK will not work 
                        # read_data[1]= '{0:08b}'.format(read_data[1])
                        # read_data[2]= '{0:08b}'.format(read_data[2])
                        # read_data[3]= '{0:08b}'.format(read_data[3])

                        # read_value = ' '.join(read_data)
                        # read_value_h = hex(db.jlink.memory_read32(reg_add,1)[0])
                        # read_value = read_value+" ( "+read_value_h+" )"
                        # db.printing(1,"**-> \t",write_add_data[0],"=> \t")

                        # # db.printing(1,"**-> \t",write_add_data[0],"=> \t",read_value,"\t<=>\t",bin(int(read_value[2:],16)))
                        # # read_value = hex(db.jlink.memory_read32(reg_add,1)[0])
                        # all_trim_reg_dict.update({write_add_data[0]:[0,read_value]})
                        db.jlink.memory_write32( int(write_add_data[0][2:],16), [0]   )
                    write_add_data = [0,0]

    # db.store_global('all_trim_regs',all_trim_reg_dict)

##################
# Meat and Potato of Code, this will start calling the functions of the program 
def parse_pfd_file(this_pfd_name):
    global VDDLO_FLAG, VDDLR_FLAG, VDDA_FLAG, IS_LR_FLAG, VDDSCR_FLAG, VGPTAT_FLAG,BG_FLAG, excelDic, SP_BLOCK_FLAG, LO_CLOCK_FLAG, write_calculated_flag, TS_CURR_TRIM_FLAG,IS_LR_BLOCK_FLAG
    global pfd_name,all_pfd,sp_dict,ts_dict,is_dict,running_report
    global look_for_subsections
    read_flag = 0
    write_flag = 0
    swd_args = []
    shorted_pads = []
    number = pfd_name.index(this_pfd_name)
    global special_script, calcdict, debugging, this_file_name,pfd_folder
    this_file_name = pfd_name[number]

    db.refresh_this_log(this_pfd_name)

    end_this_run = 0
    # db.sharing_globals(this_pfd_name,pfd_folder)
    db.store_global("pfd_name",this_pfd_name)
    db.store_global("pfd_folder",pfd_folder)
    db.store_global("check_with_expected",0)
    # while True:
    #     line = pfd_file.readline()
    #     pfd_line = line.split(" ", 1)
    #     if not line:
    #         break
    line_no = 1
    db.printing(debugging,"\n")
    print ("\n**-> ---------- Running: ",pfd_name[number],"-"*10)
    calcdict = {}
    
    list_of_pfd = [all_pfd[number]]
    name_on_progress_bar = this_file_name
    ##################### checking Expected range ####################################
    try_again = 1
    what_to_run = []
    look_for_subsections = db.get_global("check_for_subsection") 
    is_section_name_ok = 1
    while try_again:
        for line in all_pfd[number]:
            global expected_results,looking_for_expected_ranges
            if ("EXPECTED RANGE" in line):
                looking_for_expected_ranges = 1
                db.printing(1,"**-> Looking for Expected Range")
            elif (looking_for_expected_ranges ==1):
                if ("****" in line):
                    looking_for_expected_ranges = 0
                    print("\n--------------------------------------------")
                    db.store_global("check_with_expected",1)
                else:
                    if "SUBSECTIONS_PRESENT" in line:   # now redundand, as it is added in debug_options
                        look_for_subsections = 1
                        
                    else:
                        temp_range = re.split("[=\\[\\], ]",line)
                        for i in range(len(temp_range)):
                            temp_range[i] = temp_range[i].replace(" ",'')
                        # for i in temp_range:
                        #     if i =='':
                        #         temp_range.remove(i)
                        while '' in temp_range:
                            temp_range.remove('')
                        # print(temp_range)
                        try:
                            float(temp_range[1])        
                            expected_results.update({ temp_range[0] : [float(temp_range[1]),float(temp_range[2])] })
                            try:
                                expected_results.update({ temp_range[0] : [float(temp_range[1]),float(temp_range[2]),int(temp_range[3])] })
                                db.printing(1, temp_range[0]," =  \t[",float(temp_range[1]),",",float(temp_range[2]),"] \tUrgent:",int(temp_range[3]))
                            except:
                                db.printing(1, temp_range[0]," =  [",float(temp_range[1]),",",float(temp_range[2]),"]")
                                pass
                            
                        except:
                            pass
            else:
                try_again = 0

        # print("is_section_name_ok:",is_section_name_ok)
        if (look_for_subsections) :
            [what_to_run,all_section_name] = find_all_subsections_in_pfd(all_pfd[number],look_for_subsections,is_section_name_ok)
            
            if not all_section_name:    # if no subsections defned in pfd then skip
                what_to_run = []
                break
            if is_section_name_ok:
                what_to_run = db.get_global("predefined_section")
                print(what_to_run)
            for i in what_to_run:
                if i not in all_section_name:
                    db.printing(1,"**-> All subsection names do not match, pls try again (",i,")")
                    try_again = 1
                    is_section_name_ok =0
                    break
                try_again =0
            if try_again ==0:
                db.printing(1,"**-> Running subsections in order =>", what_to_run)
                look_for_subsections = 0    # completed looking for subsections and decide what to run
    global dict_of_subsections
    if what_to_run:
        list_of_pfd = []
        for i in what_to_run:
            list_of_pfd.append(dict_of_subsections[i])
    
    ### --------------   Executing pfd line by line -----------------------------
    
    what_to_run_index = 0
    for running_this_part_pfd in list_of_pfd:
        if len(list_of_pfd)>=2:
            name_on_progress_bar = what_to_run[what_to_run_index]
            what_to_run_index+=1
        with alive_bar(len(running_this_part_pfd),title=name_on_progress_bar) as bar:    
            for line in running_this_part_pfd:
                

                report = db.get_global("running_report")
                if report['1']: # send message if urgent error
                    db.report_me_i_did_something(debug_option['email'])    
                    print ("Error in ",this_pfd_name, "Line number: ",line_no)
                    print(line)
                    # icon.stop()
                    end_this_run = make_a_breakpoint(line)
                    if end_this_run==1:
                        break
                

                bar()
                if end_this_run==1:
                    break

                db.printing(debugging,line_no," ",line) # line number for ease of finding line
                line_no +=1
                pfd_line = line.split(" ",1)
                if "" == line:
                    db.printing(debugging,"**-> blank line")
                elif not line: ## redundant
                    break
#################If any special Commands in messages, add it here. 
                elif "[MSG]" == pfd_line[0]:
                    global FINALLY_VDDSCR,FINALLY_VDDA,FINALLY_VDDLO,FINALLY_VDDLR,FINALLY_IS_LR
                    # db.printing(debugging,pfd_line[1])
                    this_line = ((pfd_line[1]).split(" "))



###-------------------------------- """BASIC BLOCK CHECKS"""

#NOTE: Add code here to work with the IS LR Trimming pattern 

                    if "Start of BG" in line:
                        #binarySearchBG(swd_args[0],BG_Target)
                        BG_FLAG = 1
                    elif ("Find out the trim such that the BG Voltage" in line) and (debug_option["BG_trim_custom"]==0):
                            global BG_Target
                            BG_Target = float(line.split("==")[1].replace("V>",""))
                         



                    elif "Start of IS LR TRIM" in line:
                        IS_LR_FLAG = 1   
                    elif "End of IS LR TRIM" in line:
                        IS_LR_FLAG = 0
                        [FINALLY_IS_LR,IS_LR_VOLT] = trim_sweep(db,db.jlink,debugging,IS_LR_trimAdress,IS_LR_trim,IS_LR_BANK,IS_LR_Target,name_IS_LR)
                        IS_LR_BLOCK_FLAG = 1

                        #IS_LR_TRIM_WRITE_FLAG = 1
                        #Set Another flag here to write IS_LR_Trim value to register in the special script section
                        #see "end of SP VDD LR TRIM" as an example


                    elif "Start of Time Zero temperature measurement by using delta VBE" in line:
                        VGPTAT_FLAG = 1
                        pass
                    elif "Find out the bg_vptat trim such that VPTAT Voltage == Target_Vptat" in line:
                        VGPTAT_FLAG = 1
                        if debug_option["BG_trim_method"]==0:
                            trimlistVG = []
                            alist=[0]
                            daqReadValVG=[]
                            DAQ_ch1 = db.get_global("CH1_DAQ")  #pad_avrrdy
                            DAQ_ch2 = db.get_global("CH2_DAQ")  #pad_avren
                            b = 0
                            for i in range(0, 63):
                                findREG(db.jlink,debugging, "trim.BG_TRIM1.bg_vptat_trim", alist[0]) #will only write to the VGPTAT REGISTER
                                trimlistVG.append(b)
                                ele2 = abs(DAQread(DAQ_ch1)-DAQread(DAQ_ch2))
                                print("**-> code= ", b,", read_val=  ", ele2)
                                #pass the trim values and register readings to the registermatch file 
                                daqReadValVG.append(ele2)#read the daq reading here
                                #pass it the adress here as well VGPTAT_trimMe(trimlistVG,daqReadValVG, Target_VPtat_temp)
                                b = b + 1
                                alist[0] = b
                        # db.printing(debugging,trimlistVG)
                        # db.printing(debugging,daqReadValVG)
                        # db.printing(debugging,"THIS IS YOUR FINAL TRIM LIST MATE")
                    elif "Start of SP VDDLO TRIM" in line:
                        VDDLO_FLAG = 1
                        db.printing(debugging,"*****-> VDDDLO Flag")
                    elif "end of SP VDDLO TRIM" in line:
                        #find trim value, write into register
                        VDDLO_FLAG = 0 
                        global VDDLO_trimAdress
                        [FINALLY_VDDLO,VDDLO_volt] = trim_sweep(db,db.jlink,debugging,VDDLO_trimAdress,VDDLO_trim,VDDLO_BANK,VDDLO_Target,name_VDDLO)
                        dict.update({"VDDLO_volt":VDDLO_volt})
                        dict.update({"VDDLO_trim_code":FINALLY_VDDLO})
                        #FINALLY_VDDLO = VDDLO_trimMe(VDDLO_trimAdress,VDDLO_trim,VDDLO_BANK,VDDLO_Target,name_VDDLO)
                    elif "Start of SP VDDA TRIM" in line:
                        VDDA_FLAG = 1
                    elif "Start of SP VDD SCR TRIM" in line:
                        VDDSCR_FLAG = 1
                    elif "Start of SP VDD LR TRIM" in line:
                        VDDLR_FLAG = 1
                    elif "end of SP VDDA TRIM" in line:
                        VDDA_FLAG = 0 
                        [FINALLY_VDDA,VDDA_volt] = trim_sweep(db,db.jlink,debugging,VDDA_trimAdress,VDDA_trim,VDDA_BANK,VDDA_Target,name_VDDA) 
                        dict.update({"VDDA_volt":VDDA_volt}) 
                        dict.update({"VDDA_trim_code":FINALLY_VDDA})
                        #FINALLY_VDDA = VDDA_trimMe(VDDA_trimAdress,VDDA_trim,VDDA_BANK,VDDA_Target,name_VDDA)
                    elif "end of SP VDD SCR TRIM" in line:
                        VDDSCR_FLAG = 0
                        [FINALLY_VDDSCR,VDDSCR_volt] = trim_sweep(db,db.jlink,debugging,VDDSCR_trimAdress,VDDSCR_trim,VDDSCR_BANK,VDDSCR_TARGET,name_VDDSCR)
                        dict.update({"VDDSCR_volt":VDDSCR_volt})
                        dict.update({"VDDSCR_trim_code":FINALLY_VDDSCR})
                        #FINALLY_VDDSCR = VDDSCR_trimMe(VDDSCR_trimAdress,VDDSCR_trim,VDDSCR_BANK,VDDSCR_TARGET,name_VDDSCR)
                    elif "end of SP VDD LR TRIM" in line:
                        VDDLR_FLAG = 0
                        SP_BLOCK_FLAG = 1
                        [FINALLY_VDDLR,VDDLR_volt] = trim_sweep(db,db.jlink,debugging,VDDLR_trimAdress,VDDLR_trim,VDDLR_BANK,VDDLR_Target,name_VDDLR)
                        dict.update({"VDDLR_volt":VDDLR_volt})
                        dict.update({"VDDLR_trim_code":FINALLY_VDDLR})
                        #FINALLY_VDDLR = VDDLR_trimMe(VDDLR_trimAdress,VDDLR_trim,VDDLR_BANK,VDDLR_Target,name_VDDLR)
                    elif "Apply input reference clock" in line:
                        #fGenON()
                        LO_CLOCK_FLAG = 1



###-------------------------------- execute commands from pfd  --------------------



                    elif "<Short" in this_line:         # Specific for pi
                        if ('pad' in this_line[1]) and ('pad' in this_line[3]): 
                            shorted_pads = [this_line[1],this_line[3].replace('>','')]
                            db.printing(debugging,"**-> Saving pad names to apply something on both, on next PAD message")
                            

                    elif "PAD" in pfd_line[1]:
                        # if shorted_pads:
                        #     apply_something_for_shorted_pads(shorted_pads,pfd_line[1])
                        #     # send_ate_command(pfd_line[1])
                        # else:
                            send_ate_command(pfd_line[1],shorted_pads)
                            shorted_pads=[]

                    

                
                    elif "<COMPUTE:" == this_line[0] or "<COMPUTE_AFTER:" == this_line[0]:
                        parse_and_store_equation(line,calcdict,db.jlink) 
                    elif "<End of" in line:
                        reset_and_store_values(line,pfd_name[number])  
                        db.printing(debugging,"**-> Next Test") 
                    elif "Run binary search on ts_tsidac_cur_trim" in line:
                        TS_CURR_TRIM_FLAG = 1
                        global TS_current_target
                        if debug_option["TS_trim_custom"] == 0:
                            TS_current_target = line.split(" ")[-1].replace('uA>',"")

                    elif ("***" in line):# and ("Start" in line):
                        print(line.replace("[MSG]","\t"))
                        report = db.get_global("running_report")
                        report['0'].append(line.replace("[MSG]","\t"))
                        db.store_global("running_report",report)
                                    
                    else:
                        db.printing(debugging,"**-> Doing nothing")
            
                elif "[LINE_RESET]" == pfd_line[0]:
                    if jlink_ENABLE:
                        db.printing(debugging,"3")
                        db.jlink.reset()
                    # Skipping LINE_RESET commands since it is performed in the code above
                    for _ in range(5):
                        next(all_pfd)

                elif "[WAIT]" == pfd_line[0]:
                    delay = clean_wait_args(pfd_line[1])
                    db.printing(debugging,"**->Delay in sec:",delay)
                    sleep(delay)
                    # sleep(0.5)

                elif "[WRDP]" == pfd_line[0] or "[WRAP]" == pfd_line[0]:
                    write_flag += 1
                    write_args = pfd_line[1]
                    write_args = write_args.split()
                    global special_script
                    [write_args,special_script] = clean_swd_args(write_args,special_script)
                    swd_args.append(write_args[1])

                    if write_flag == 2 and read_flag == 0:#IMPORTANT, dont think you need to have name_,
                        #Special _Script flag is set when the write trim command is seen in the PFD 
                        # db.printing(debugging,special_script)
                        # db.printing(debugging,"hui")
                        if special_script == 1:
                            #NOTE: REWORK THIS
                            #Grab register specific register name 
                            #grab the value that you want to write
                            #write into that value 
                            #binarySearchBG(swd_args[0],BG_Target)
                            if SP_BLOCK_FLAG ==1:
                                findREG(db.jlink,debugging,name_VDDLO,FINALLY_VDDLO)
                                findREG(db.jlink,debugging,name_VDDA,FINALLY_VDDA)
                                findREG(db.jlink,debugging,name_VDDLR,FINALLY_VDDLR)
                                findREG(db.jlink,debugging,name_VDDSCR,FINALLY_VDDSCR)
                                SP_BLOCK_FLAG = 0
                            if IS_LR_BLOCK_FLAG ==1:
                                findREG(db.jlink,debugging,name_IS_LR,FINALLY_IS_LR)
                                IS_LR_BLOCK_FLAG = 0
                                
                            if BG_FLAG ==1:
                                #swd_args[0] should already have the 0x70000000 adress, dont need to do code below 
                                swd_args[0] = returnAdress(db.jlink,debugging,"trim.BG_TRIM1.bg_trim")
                                # SweepSearchBG(db.jlink,debugging,swd_args[0],BG_Target)
                                dict.update({"BG_Target":BG_Target})
                                if debug_option["BG_trim_method"]==0:
                                    [code,val] = SweepSearchBG(db.jlink,debugging,"trim.BG_TRIM1.bg_trim",BG_Target)
                                else:
                                    [code,val] = BinarySearchBG(db.jlink,debugging,"trim.BG_TRIM1.bg_trim",BG_Target)
                                dict.update({"BG_Final":code})
                                dict.update({"BG_Final_val":val})
                                BG_FLAG =0
                            if VGPTAT_FLAG ==1:
                                if debug_option["BG_trim_method"]==0:
                                    [FINALLY_VGPTAT,VGPTAT] = VGPTAT_trimMe(db,db.jlink,debugging,trimlistVG,daqReadValVG, Target_VPtat_temp)
                                    
                                #findREG("trim.BG_TRIM1.bg_trim",BG_trim_value)
                                else:
                                    [FINALLY_VGPTAT,VGPTAT]  = VGPTAT_binary(db.jlink,debugging,Target_VPtat_temp,"trim.BG_TRIM1.bg_vptat_trim")
                                db.store_global("VGPTAT_val",VGPTAT)
                                db.store_global("VGPTAT_code",FINALLY_VGPTAT)
                                dict.update({"VGPTAT_val":VGPTAT})
                                dict.update({"VGPTAT_code":FINALLY_VGPTAT})
                                findREG(db.jlink,debugging,"trim.BG_TRIM1.bg_vptat_trim",FINALLY_VGPTAT)
                                VGPTAT_FLAG = 0
                                #not necessary, but can use findREG, double check 
                                #findREG("BG registname here",FINALLY_VDDSCR)
                            if LO_CLOCK_FLAG ==1:
                                #fGenOff()
                                lo_clk_trim_coarse = newSave["teststat.BIST_LO2.lo_clk_trim_coarse_bist"]
                                db.printing(debugging,"Your lo_clk_trim_coarse is: ", lo_clk_trim_coarse)
                                lo_clk_trim_fine = newSave["teststat.BIST_LO2.lo_clk_trim_fine_bist"]
                                db.printing(debugging,"Your lo_clk_trim_fine is: ", lo_clk_trim_fine)
                                #lo_clk_trim_coarse_20mhz = newSave[name_lo_clk_trim_coarse_20mhz]
                                #lo_ptat_trim = newSave [name_lo_ptat_trim]
                                findREG(db.jlink,debugging,name_lo_clk_trim_coarse,lo_clk_trim_coarse)
                                findREG(db.jlink,debugging,name_lo_clk_trim_fine,lo_clk_trim_fine)
                                sp_dict.update({name_lo_clk_trim_coarse: lo_clk_trim_coarse})
                                sp_dict.update({name_lo_clk_trim_fine: lo_clk_trim_fine})
                                swd_args[0] = 1879048200
                                read_swd_simple(swd_args)
                                LO_CLOCK_FLAG = 0
                                # break
                            if write_calculated_flag ==1:
                                swd_args[1] = calc_for_write
                                
                                # db.printing(debugging,swd_args)
                                # write_full_reg(swd_args[0],swd_args[1],db.jlink)
                                db.printing(debugging,"**-> Write: Addr: %s, Data: %s" %(hex(swd_args[0]), hex(swd_args[1])))
                                db.jlink.memory_write32(swd_args[0],[swd_args[1]])
                                write_calculated_flag = 0
                            if TS_CURR_TRIM_FLAG ==1:
                                # global TS_current_target
                                ts_code = ts_current_trim(db,db.jlink,debugging,swd_args[0],TS_current_target,debug_option["TS_trim_method"],"trim.TS_TRIM6.ts_tsidac_cur_trim")
                                calcdict.update({"TS_Trim_Code":ts_code[0]})      
                                calcdict.update({"TS_Trim_Current":float(ts_code[1])})   
                                print("-> TS_Trim_Code:",ts_code[0],"\n-> TS_Trim_Current: -> ",float(ts_code[1]))     
                                
                            # if (write_calculated_flag ==0):
                            saveAddress = swd_args[0]       
                            # db.printing(debugging,saveAddress)
                            special_script = 0
                            read_flag = 0
                            write_flag = 0
                            swd_args = []
                            global readCommentReg
                            readCommentReg = 1
                        else:
                            write_swd(db.jlink,swd_args[0:2])
                            read_flag = 0
                            write_flag = 0
                            swd_args = []       

                elif "[RDDP]" == pfd_line[0] or "[RDAP]" == pfd_line[0]:
                    read_flag += 1
                    if "[RDDP]" == pfd_line[0]:
                        read_args = pfd_line[1]
                        read_args = read_args.split()
                        read_args = clean_swd_args(read_args,special_script)[0]
                        # swd_args.append(read_args[0])
                        swd_args.append(read_args[1])

                    if write_flag == 1 and read_flag == 2: #need to add a special command here for BIST READINGS
                        saveAddress = swd_args[0]
                        sleep(0.1)
                        read_swd(swd_args,regcapture,saveMask) #had it as read_swd(swd_args,hi)
                        db.printing(debugging,"**-> Mask:\t",swd_args[1])
                        read_flag = 0
                        write_flag = 0
                        swd_args = []
                elif "[TEST]" == pfd_line[0]:
                    for i in calcdict:
                        try:
                            db.printing(debugging,i,"-"*(50-len(i)),"->",calcdict[i])
                        except:
                            pass
                    print(line)
                    # breakpt = input("*"*60,"\n**-> Waiting at breakpoint [TEST], input anything to cont.")

                    end_this_run = make_a_breakpoint(line)
                    if end_this_run==1:
                        break

                else:


                    #NOTE: IF reg_capture in line, capture name and mask value 

                    #code below will extract register name to write into if the special script command flag is active
                    #else it will just printing out the line 
                    global counter
                    global readCounter 
                    lookForReadReg = line
                    '''
                    global expected_results,looking_for_expected_ranges
                    if ("EXPECTED RANGE" in line):
                        looking_for_expected_ranges = 1
                        db.printing(1,"**-> Looking for Expected Range")
                    elif (looking_for_expected_ranges ==1):
                        if ("****" in line):
                            looking_for_expected_ranges = 0
                            print("\n--------------------------------------------")
                            db.store_global("check_with_expected",1)
                        else:
                            if "SUBSECTIONS_PRESENT" in line:
                                look_for_subsections = 1
                                
                            else:
                                temp_range = re.split("[=\\[\\], ]",line)
                                for i in range(len(temp_range)):
                                    temp_range[i] = temp_range[i].replace(" ",'')
                                # for i in temp_range:
                                #     if i =='':
                                #         temp_range.remove(i)
                                while '' in temp_range:
                                    temp_range.remove('')
                                # print(temp_range)
                                try:
                                    float(temp_range[1])        
                                    expected_results.update({ temp_range[0] : [float(temp_range[1]),float(temp_range[2])] })
                                    try:
                                        expected_results.update({ temp_range[0] : [float(temp_range[1]),float(temp_range[2]),int(temp_range[3])] })
                                        db.printing(1, temp_range[0]," =  \t[",float(temp_range[1]),",",float(temp_range[2]),"] \tUrgent:",int(temp_range[3]))
                                    except:
                                        db.printing(1, temp_range[0]," =  [",float(temp_range[1]),",",float(temp_range[2]),"]")
                                        pass
                                    
                                except:
                                    pass
                    '''                 

                    if "reg_capture " in lookForReadReg: 
                        global dataRegisterName
                        saveMask.clear() #clearing mask to ensure register read 
                        #may need to comment out code above 
                        start = ':'
                        end = ''
                        start1 = 'capture'
                        end1 = 'mask'
                        s = line
                        g = line
                        s = (s[s.find(start) + len(start):s.rfind(end)])     #This will save the masking number
                        g = (g[g.find(start1) + len(start1):g.rfind(end1)])  #This will save the name of the register
                        g = g.replace(" ","") #This will get rid of the spaces in the register name
                        s = s.replace("\n","") #NEED TO GET RID OF THE \n IN THE CODE!
                        db.printing(debugging,"**-> Storing Mask: ", s)
                        regcapture.append(g)
                        dataRegisterName.append(g)
                        saveMask.update({g: s}) #this will be filled with read_capture registers


                    # elif ("read reg" or "write reg" in lookForReadReg):
                    #     pass

                        #same line as reg_capture, and write reg is not used
                        #This is where
                    elif (line.find("reg_write_computed_")!=-1) and ("sp" not in pfd_name[number]) and (TS_CURR_TRIM_FLAG!=1):
                        [calc_for_write,write_calculated_flag] = find_and_save_computed_value(line,calcdict,db.jlink)

                    
                    else:
                        db.printing(debugging,"**-> Unused/Comment Line")

                db.printing(debugging,"________________________________")


                if(db.get_global("check_with_expected") == 1):  # check with range at start of pfd
                    end_this_run = check_results_with_expected()
                if end_this_run==1:
                    break

    global is_accuracy_dict,pi_dict
    if ("vs.pfd" in pfd_name[number] ):# or ("ate_1_test" in pfd_name[number] ):
        print_results_for_vs()
    elif("sp" in pfd_name[number]):
        sp_dict = dict
    elif("_ts" in pfd_name[number]):
        ts_dict = calcdict
    elif("_is" in pfd_name[number]):
        if ('accuracy' in pfd_name[number]):
            is_accuracy_dict = calcdict
        else:
            is_dict = calcdict
        # print("-"*35," All Results ","-"*35)
        if debugging == 1:
            for i in is_dict:
                print(i,"-"*(50-len(i))," -> ",is_dict[i])
    elif ("pi.pfd" in pfd_name[number]):
        # print(calcdict)
        pi_dict = calcdict
        
        
                
        
    db.printing(debugging,"___________________")
    # global running_report
    this_complete = "Completed running-> "+this_pfd_name
    report = db.get_global("running_report")
    report['0'].append(this_complete)
    db.store_global("running_report",report)

    save_in_excel() 
    reset_all_dict()
    # turn_off_smu()










#############
all_checks_complete = {}
def check_results_with_expected(): 
    global expected_results
    global calcdict,dict
    all_checks = {}
    all_checks.update(calcdict)
    all_checks.update(dict)
    end_this_run = 0

    for i in all_checks_complete: # remove already checked value 
        all_checks.pop(i)
    
    for i in expected_results:
        if i in all_checks:
            if (all_checks[i] - expected_results[i][1] >= 0.0 ) or (all_checks[i] - expected_results[i][0] <= 0.0):
                db.printing(1,"**-> ",i," is out of expected range, \nExpected:[",expected_results[i][0],",",expected_results[i][1],"]\nObserved: ",all_checks[i]) 
                report = db.get_global("running_report")
                temp = str(i)+" is out of expected range, \n\t\tExpected:["+str(expected_results[i][0])+","+str(expected_results[i][1])+"]\n\t\tObserved: "+str(all_checks[i])
                report['1'].append(temp)
                db.store_global("running_report",report)
                db.report_me_i_did_something(debug_option['email']) 
                report['1'].remove(temp)    # So that email is not continiously sent
                db.store_global("running_report",report)

                if (len(expected_results[i]) >=3 ):
                    if (expected_results[i][2] == 1):
                        end_this_run = make_a_breakpoint('')
            
            all_checks_complete.update({i:all_checks[i]})
            # del all_checks[i]
    return end_this_run



def print_results_for_vs(): ## print in terminal, can be skipped now as now we are printing results as we calculate them and this will just reprint all results in terminal
    global all_stored_vs, all_stored_vs_Tests
    print("\n\n ------ Printing again for simplicity ---->>")
    for  j in range(len(all_stored_vs)):
        print ("-"*int((60-len(all_stored_vs_Tests[j]))/2),all_stored_vs_Tests[j],"-"*int((60-len(all_stored_vs_Tests[j]))/2))
        for i in sorted (all_stored_vs[j].keys()) :
            this_value = all_stored_vs[j][i]
            if '0x' in str(this_value):
                # db.printing(debugging,this_value)
                # db.printing(debugging,"i=",i)
                print (i,"-"*(43-len(i)), "----=>",this_value, "\t <=> ", bin(int(this_value,16)))
            else:
                if ("%" in i):
                    this_value = str(this_value)+"uV"
                print (i,"-"*(43-len(i)), "----=>",this_value)



def reset_and_store_values(line,this_pfd_name):
    global calcdict, all_stored_vs, all_stored_vs_Tests,vs_adc_dict,vs_adc_x
    global VS_Gain_Trim, VS_Digital_Offset_BIST, VS_Analog_Offset_BIST,VS_Gain_Trim_Retest,VS_Offset_BIST_retest,VOS_DAC_offset_DNL_test,DNL_Test
    if ("vs.pfd" in this_pfd_name):
        if "VS Gain Trim>" in line:
            VS_Gain_Trim = calcdict
        elif "VS Digital Offset BIST>" in line:
            VS_Digital_Offset_BIST = calcdict
        elif "VS Analog (VOS_DAC) Offset BIST>" in line:
            VS_Analog_Offset_BIST = calcdict   
        elif "VS Gain Trim Retest>" in line:
            VS_Gain_Trim_Retest = calcdict
        elif "VS Offset BIST retest>" in line:
            VS_Offset_BIST_retest = calcdict
        elif "VOS_DAC 0V offset trim for DNL test>" in line:
            # for data in calcdict:
                # calcdict[data+"_offs_DNL"] = calcdict.pop(data)
            VOS_DAC_offset_DNL_test = {k+"_offs_DNL":v for k,v in calcdict.items()}
        elif "<End of DNL test>" in line:
            # for data in calcdict:
                # calcdict[data+"_DNL"] = calcdict.pop(data)
            DNL_Test = {k+"_DNL":v for k,v in calcdict.items()}

        all_stored_vs = [VS_Gain_Trim,VS_Digital_Offset_BIST,VS_Analog_Offset_BIST,VS_Gain_Trim_Retest,VS_Offset_BIST_retest,VOS_DAC_offset_DNL_test,DNL_Test]
        all_stored_vs_Tests = ["VS_Gain_Trim","VS_Digital_Offset_BIST","VS_Analog_Offset_BIST","VS_Gain_Trim_Retest","VS_Offset_BIST_retest","VOS_DAC_offset_DNL_test","DNL_Test"]
        calcdict = {}

    elif ("vs_adc" in this_pfd_name):
        vs_adc_dict = calcdict

        for i in vs_adc_dict:
            vs_adc_x.append(float(i[len(i)-9:]))
    
            
   

def reset_all_dict():
    global dict, sp_dict, all_stored_vs, vs_adc_dict, ts_dict ,is_dict ,is_accuracy_dict, pi_dict
    dict_list = [[dict],[sp_dict],all_stored_vs,[vs_adc_dict],[ts_dict],[is_dict],[is_accuracy_dict],[is_accuracy_dict],[pi_dict]] 
    dict = {}
    sp_dict = {}
    vs_adc_dict = {}
    ts_dict = {}
    is_accuracy_dict = {}
    is_dict = {}
    pi_dict = {}
    for i in range(len(all_stored_vs)):
        all_stored_vs[i] = {}


def Read_VS_volt_from_pfd(Board_addr,accuracy_test_pfd):
    reset_all_dict()    # resetting previous stored data
    vs_temperature_dict = {}
    db.store_global("pfd_name",accuracy_test_pfd)
    # print(db.jlink.connected())
    parse_pfd_file(accuracy_test_pfd)
    vs_temperature_dict = calcdict
    for i in vs_temperature_dict:
        # print(i,"----",vs_temperature_dict[i])
        this_val = int(bin(int(vs_temperature_dict[i]) + 2**13)[2:],2) # (this_val/8)*0.00125
        # print(i,"----",this_val)
        this_val = (this_val/8)*0.00125
        # print(i,"----",this_val)
        vs_temperature_dict.update({i:this_val})
        # print(i,"\t\t",vs_temperature_dict[i])
    db.store_global("VS_temperature_dict",vs_temperature_dict)
    # print("Parsing vs accuracy test")  


def Read_IS_volt_from_pfd(Board_addr,IS_1K_pfd):
    reset_all_dict()    # resetting previous stored data
    db.store_global("pfd_name",IS_1K_pfd)
    is_temperature_dict = {}
    parse_pfd_file(IS_1K_pfd)
    is_temperature_dict = calcdict
    db.store_global("IS_temperature_dict",is_temperature_dict)


def Read_VS_volt(Board_addr,voltage_range):
    add_data = ['0x70004858','0x000000f4']
    db.jlink.memory_write32(int(add_data[0].replace('0x',''),16), [int(add_data[1].replace('0x',''),16)])
    add_data = ['0x7000480c','0x00000003']
    db.jlink.memory_write32(int(add_data[0].replace('0x',''),16), [int(add_data[1].replace('0x',''),16)])
    add_data = ['0x7000484c','0x7ed0333c']
    db.jlink.memory_write32(int(add_data[0].replace('0x',''),16), [int(add_data[1].replace('0x',''),16)])
    vs_temperature_dict = {}
    for i in voltage_range:

        if ( float(i) - 0.95 < 0):  # VID select
            add_data = ['0x70005C3C','0x0']
            db.jlink.memory_write32(int(add_data[0].replace('0x',''),16), [int(add_data[1].replace('0x',''),16)])
            add_data = ['0x7000583C','0x0']
            db.jlink.memory_write32(int(add_data[0].replace('0x',''),16), [int(add_data[1].replace('0x',''),16)])
        elif ( float(i) - 1.7 < 0):
            add_data = ['0x70005C3C','0x2']
            db.jlink.memory_write32(int(add_data[0].replace('0x',''),16), [int(add_data[1].replace('0x',''),16)])
            add_data = ['0x7000583C','0x2']
            db.jlink.memory_write32(int(add_data[0].replace('0x',''),16), [int(add_data[1].replace('0x',''),16)])
        else:
            add_data = ['0x70005C3C','0x4']
            db.jlink.memory_write32(int(add_data[0].replace('0x',''),16), [int(add_data[1].replace('0x',''),16)])
            add_data = ['0x7000583C','0x4']
            db.jlink.memory_write32(int(add_data[0].replace('0x',''),16), [int(add_data[1].replace('0x',''),16)])

        set_voltage_or_current(debugging,(str(i)+'V'),"pad_vsenp")
        set_voltage_or_current(debugging,(str(i)+'V'),"pad_bvsenp")
        
        add_data = ['0x70004850','0x00003a1e']
        db.jlink.memory_write32(int(add_data[0].replace('0x',''),16), [int(add_data[1].replace('0x',''),16)])
        sleep(0.01)
        name0  = 'vadc_ch0_avg_'+str(i)
        read0 = read_reg_data(db.jlink,debugging,"teststat.BIST_AVG10.vadc_ch0_avg")
        read0 = (read0/8)*0.00125

        sleep(0.01)
        name1  = 'vadc_ch1_avg_'+str(i)
        read1 = read_reg_data(db.jlink,debugging,"teststat.BIST_AVG11.vadc_ch1_avg")
        read1 = (read1/8)*0.00125  

        db.printing(1,name0,"-"*(35-len(name0)),read0) 
        db.printing(1,name1,"-"*(35-len(name1)),read1)
        vs_temperature_dict.update({name0:read0})
        vs_temperature_dict.update({name1:read1})
        add_data = ['0x70004850','0x0000021e']
        db.jlink.memory_write32(int(add_data[0].replace('0x',''),16), [int(add_data[1].replace('0x',''),16)])
    db.store_global("VS_temperature_dict",vs_temperature_dict)
        
    return 1


# {"BG_Reset":'-',"BG_60_per":'-', "BG_40_per":'-',"BG_Final":'-','BG_Target':'-'}
xl_dict = [{'BG_Target':'-','BG_T_code':'-',"BG_Final":'-','VPTAT_Final':'-'}, 
{"tmux_diff_sp_vddlo_code_all":"-","tmux_diff_sp_vdda_code_all":"-","tmux_diff_sp_vdd_scr_code_all":"-","tmux_diff_sp_vdd_lr_code_all":"-","tmux_diff_sp_vddlo_code_final":"-","tmux_sp_vdda_code_final":"-","tmux_sp_vdd_scr_code_final":"-","tmux_sp_vdd_lr_code_final":"-","VDDLO_trim_code":"-","VDDA_trim_code":"-","VDDSCR_trim_code":"-","VDDLR_trim_code":"-"},
# {"teststat.BIST_VS_GAIN.vs0_test_gain_comp":"-","teststat.BIST_VS_GAIN.vs1_test_gain_comp":"-","vsp0_adc_gain_vid1":"-","vsp1_adc_gain_vid1":"-","vsp0_adc_gain_vid2":"-","vsp1_adc_gain_vid2":"-","vsp0_adc_gain_vid3":"-","vsp1_adc_gain_vid3":"-","teststat.BIST_VS_OFFS.vs0_test_offs":"-","teststat.BIST_VS_OFFS.vs1_test_offs":"-","vsp0_adc_offset_vid1":"-","vsp1_adc_offset_vid1":"-","vsp0_adc_offset_vid2":"-","vsp1_adc_offset_vid2":"-","vsp0_adc_offset_vid3":"-","vsp1_adc_offset_vid3":"-","teststat.BIST_AVG10.vs0_os_trim_tm":"-","teststat.BIST_AVG11.vs1_os_trim_tm":"-","vs0_os_trim_bif_ana":"-","vs1_os_trim_bif_ana":"-","vs0_os_trim_bif":"-","vs1_os_trim_bif":"-","vs0_vid1_test_gain_comp_retest":"-","vs0_vid1_test_gain_comp_retest_gain_error_%":"-","vs1_vid1_test_gain_comp_retest":"-","vs1_vid1_test_gain_comp_retest_gain_error_%":"-","vs0_vid1_test_gain_err_retest":"-","vs1_vid1_test_gain_err_retest":"-","teststat.BIST_VS_GAIN.vs0_test_gain_comp":"-","teststat.BIST_VS_GAIN.vs1_test_gain_comp":"-","vs0_vid2_test_gain_comp_retest":"-","vs0_vid2_test_gain_comp_retest_gain_error_%":"-","vs1_vid2_test_gain_comp_retest":"-","vs1_vid2_test_gain_comp_retest_gain_error_%":"-","vs0_vid2_test_gain_err_retest":"-","vs1_vid2_test_gain_err_retest":"-","vs0_vid3_test_gain_comp_retest":"-","vs0_vid3_test_gain_comp_retest_gain_error_%":"-","vs1_vid3_test_gain_comp_retest":"-","vs1_vid3_test_gain_comp_retest_gain_error_%":"-","vs0_vid3_test_gain_err_retest":"-","vs1_vid3_test_gain_err_retest":"-","vs0_vid1_test_offs_retest":"-","vs0_vid1_test_offs_retest_error_mV":"-","vs1_vid1_test_offs_retest":"-","vs1_vid1_test_offs_retest_error_mV":"-","vs0_vid1_test_offs_err_retest":"-","vs1_vid1_test_offs_err_retest":"-","teststat.BIST_VS_OFFS.vs0_test_offs":"-","teststat.BIST_VS_OFFS.vs1_test_offs":"-","vs0_vid2_test_offs_retest":"-","vs0_vid2_test_offs_retest_error_mV":"-","vs1_vid2_test_offs_retest":"-","vs1_vid2_test_offs_retest_error_mV":"-","vs0_vid2_test_offs_err_retest":"-","vs1_vid2_test_offs_err_retest":"-","vs0_vid3_test_offs_retest":"-","vs0_vid3_test_offs_retest_error_mV":"-","vs1_vid3_test_offs_retest":"-","vs1_vid3_test_offs_retest_error_mV":"-","vs0_vid3_test_offs_err_retest":"-","vs1_vid3_test_offs_err_retest":"-","teststat.BIST_AVG10.vs0_os_trim_avg_offs_DNL":"-","teststat.BIST_AVG11.vs1_os_trim_avg_offs_DNL":"-","vs0_os_trim_bif_offs_DNL":"-","vs0_os_trim_bif_mV_offs_DNL":"-","vs1_os_trim_bif_offs_DNL":"-","vs1_os_trim_bif_mV_offs_DNL":"-","vs0_test_dnl_DNL":"-","vs1_test_dnl_DNL":"-","teststat.BIST_VS_DNL.vs0_test_dnl_DNL":"-","teststat.BIST_VS_DNL.vs1_test_dnl_DNL":"-","vs0_os_trim_bif_DNL":"-","vs1_os_trim_bif_DNL":"-"}
{"vsp0_adc_gain_vid1":"-","vsp1_adc_gain_vid1":"-","vsp0_adc_gain_vid2":"-","vsp1_adc_gain_vid2":"-","vsp0_adc_gain_vid3":"-","vsp1_adc_gain_vid3":"-","vsp0_adc_offset_vid1":"-","vsp1_adc_offset_vid1":"-","vsp0_adc_offset_vid2":"-","vsp1_adc_offset_vid2":"-","vsp0_adc_offset_vid3":"-","vsp1_adc_offset_vid3":"-","vs0_os_trim_bif_ana":"-","vs1_os_trim_bif_ana":"-","vs0_os_trim_bif":"-","vs1_os_trim_bif":"-","vs0_vid1_test_gain_comp_retest":"-","vs0_vid1_test_gain_comp_retest_gain_error_%":"-","vs1_vid1_test_gain_comp_retest":"-","vs1_vid1_test_gain_comp_retest_gain_error_%":"-","vs0_vid1_test_gain_err_retest":"-","vs1_vid1_test_gain_err_retest":"-","vs0_vid2_test_gain_comp_retest":"-","vs0_vid2_test_gain_comp_retest_gain_error_%":"-","vs1_vid2_test_gain_comp_retest":"-","vs1_vid2_test_gain_comp_retest_gain_error_%":"-","vs0_vid2_test_gain_err_retest":"-","vs1_vid2_test_gain_err_retest":"-","vs0_vid3_test_gain_comp_retest":"-","vs0_vid3_test_gain_comp_retest_gain_error_%":"-","vs1_vid3_test_gain_comp_retest":"-","vs1_vid3_test_gain_comp_retest_gain_error_%":"-","vs0_vid3_test_gain_err_retest":"-","vs1_vid3_test_gain_err_retest":"-","vs0_vid1_test_offs_retest":"-","vs0_vid1_test_offs_retest_error_mV":"-","vs1_vid1_test_offs_retest":"-","vs1_vid1_test_offs_retest_error_mV":"-","vs0_vid1_test_offs_err_retest":"-","vs1_vid1_test_offs_err_retest":"-","vs0_vid2_test_offs_retest":"-","vs0_vid2_test_offs_retest_error_mV":"-","vs1_vid2_test_offs_retest":"-","vs1_vid2_test_offs_retest_error_mV":"-","vs0_vid2_test_offs_err_retest":"-","vs1_vid2_test_offs_err_retest":"-","vs0_vid3_test_offs_retest":"-","vs0_vid3_test_offs_retest_error_mV":"-","vs1_vid3_test_offs_retest":"-","vs1_vid3_test_offs_retest_error_mV":"-","vs0_vid3_test_offs_err_retest":"-","vs1_vid3_test_offs_err_retest":"-","vs0_os_trim_bif_offs_DNL":"-","vs0_os_trim_bif_mV_offs_DNL":"-","vs1_os_trim_bif_offs_DNL":"-","vs1_os_trim_bif_mV_offs_DNL":"-","vs0_test_dnl_DNL":"-","vs1_test_dnl_DNL":"-","vs0_os_trim_bif_DNL":"-","vs1_os_trim_bif_DNL":"-"}
,{}# vs_adc
,{"TS_Trim_Code":"-","TS_Trim_Current":"-","tsen_adc_src4_ate_gain":"-","tsen_adc_src5_ate_gain":"-","tsen_adc_src6_ate_gain":"-","tsen_adc_src7_ate_gain":"-","tsen_adc_src4_ate_offset":"-","tsen_adc_src5_ate_offset":"-","tsen_adc_src6_ate_offset":"-","tsen_adc_src7_ate_offset":"-"}
,{}# is
,{},{}# is_accuracy, is_accuracy_error
,{}# PI
]
xl_sheet = ["BG","SP","VS","VS_ADC","TS","IS","IS_1K_all","IS_1K","PI"]
xl_temp_sheet = ["Trim_Codes","Trim_error","BG_TEMP","BG_TEMP_err","SP_TEMP","VS_TEMP","VS_TEMP_err","SP_TEMP_err","IS_TEMP"]
def save_in_excel():
    # turn_off_smu()

    global dict,pfd_folder, all_stored_vs,vs_adc_dict,vs_adc_x,xl_dict,ts_dict,pi_dict
    
    
    report = db.get_global("running_report")
    if report['1']: # send message if urgent error
        db.report_me_i_did_something(debug_option['email'])    
        # print ("Error in ",this_pfd_name, "Line number: ",line_no)
        # print(line)
        # icon.stop()
        end_this_run = make_a_breakpoint('')
        if end_this_run==1:
            return
    
    ic_serial_number = read_reg_data(db.jlink,0,"trim.IC_SERIAL.ic_serial")

    # ic_serial_number = '.'
    this_dir = os.getcwd()
    pfd_folder = db.get_global("pfd_folder")
    if (pfd_folder not in this_dir):
        temp = this_dir+"\\"+pfd_folder
        os.chdir(temp)
    # print(os.getcwd())
    excel_file = db.get_global("excel_file")

    # 
    

    # print("**-> Results in: ",excel_file)
    wb = load_workbook(excel_file)

    # IF RUNNING FOR FIRST TIME, create results.xlsx with sheets names as xl_sheet+xl_temp_sheet
    temp = xl_sheet+xl_temp_sheet
    # print(temp)

    for i in temp:  # create sheets if not present
        wb_sheets = wb.sheetnames
        if i not in wb_sheets:
            wb.create_sheet(i)



    T_now = datetime.now()
    #  EDIT HERE FOR NEW PFD -----
    try:
        BG_Final = db.get_global("BG_Final")
        dict.update({"BG_Final":BG_Final})
        BG_trim_code = db.get_global("BG_trim_code")
        dict.update({"BG_T_code":BG_trim_code})
    except:
        pass

    d_list = [[dict],[sp_dict],all_stored_vs,[vs_adc_dict],[ts_dict],[is_dict],[is_accuracy_dict],[is_accuracy_dict],[pi_dict]] # same order as xl_sheet

    for i in range(len(xl_sheet)):
        # print(xl_sheet[i])
        # print((d_list[i]))
        
        #----dont save if pfd not run ----- continue to next-----
        if (bool(d_list[i])==0): # for empth vs
            continue
        if (bool(d_list[i][0])==0):# or (not xl_dict[i]) :
            continue    # empty
        if (bool(xl_dict[i])==1) and (bool(d_list[i][0])==1):
            if list(xl_dict[i])[0] not in d_list[i][0]: 
                continue    # not empty but we dont want to print, bcz we didnt run this pfd this time.

        # --------- update xl data ----------
        if (d_list[i][0][list(d_list[i][0])[0]] != '-'):
            print("**-> Saving-> ",xl_sheet[i])
            xl = wb[xl_sheet[i]]
            
            start_row = 7
            start_col = 0
            
            col = start_col + 2       
            for dictionary in d_list[i]:
                for data in dictionary:
                    """
                    # update this in xl_dict  during setup to uniformalize which columns need to be written in xl
                    run vvvv print line
                    """
                    # print('"',data,'":\"-\"', end=',',sep="")     
                    if data in xl_dict[i]:
                        xl_dict[i].update({data:dictionary[data]})
                        # print(data,"-"*(43-len(data)), "----=>",dictionary[data]) # TO PRINT all values to be saved on terminal

            # ********* Save everything, if not then add what all to save in the list above this function**********
            if ("VS_ADC" == xl_sheet[i]) or ("VS_TEMP" == xl_sheet[i]) :
                xl_dict[i]=collections.OrderedDict(sorted(vs_adc_dict.items())) # sort by name/channel so its easier to read
                xl.cell(row=start_row-2,column=start_col+1).value = "Latest test-error mV"
                xl.cell(row=start_row-3,column=start_col+1).value = "Ideal Value"
            
            elif ("IS" == xl_sheet[i]):
            #     xl_dict[i]=collections.OrderedDict(sorted(is_dict.items()))
                xl_dict[i]=is_dict
            elif ("IS_1K_all" == xl_sheet[i]):
                xl_dict[i] = is_accuracy_dict
            elif ("IS_1K" == xl_sheet[i]):
                xl_dict[i] = is_accuracy_dict #collections.OrderedDict(sorted(is_accuracy_dict.items()))
            elif ("PI" ==xl_sheet[i]):
                xl_dict[i] = pi_dict
            
            
            is_it_rows = 0
            if ("ADC" in xl_sheet[i]) or( ('TEMP') in xl_sheet[i]):
                is_it_rows = 1
            else:
                is_it_rows= 0

            if (is_it_rows):    # can be merged in above if, but kept here to check/move this below
                xl.insert_rows(idx=start_row+1) # insert blank row at line = idx
            else:
                xl.insert_cols(idx=start_col+2) # insert blank row at line = idx


            if ("IS_1K" != xl_sheet[i]):
                
                j=0
                k=0
                for data in xl_dict[i]: # write content in a row
                    
                    
                    if ("VS_ADC" == xl_sheet[i]) or ("VS_TEMP" == xl_sheet[i]):
                        is_it_rows = 1
                        xl.cell(row=start_row,column=col).value = data
                        xl.cell(row=start_row,column=col).font = Font(color='00000000', bold=True)
                        if 2*j >= len(vs_adc_x):
                            k=j
                            j=0
                        xl.cell(row=start_row-3,column=col).value = vs_adc_x[2*j]
                        # this_val = int(bin((xl_dict[i][data]) + 2**13)[2:],2) 
                        this_val = xl_dict[i][data]
                        # print(xl_dict[i][data])
                        # print(this_val)
                        # print((this_val/8)*0.00125)
                        xl.cell(row=start_row+1,column=col).value = (this_val/8)*0.00125
                        # xl.cell(row=start_row-2,column=col).value = -1*((float(vs_adc_x[2*j])-(this_val/8)*0.00125)/float(vs_adc_x[2*j]))*100
                        # xl.cell(row=start_row-2,column=col).value = (((float(vs_adc_x[2*j])-(this_val/8)*0.00125)))*1000  # error in mV
                        xl.cell(row=start_row-2,column=col).value = ( float(vs_adc_x[2*j]) - (this_val/8)*0.00125 )*1000  # error in mV
                        j+=1
                        
                    else:
                        is_it_rows = 0
                        xl.cell(column=start_col+1,row=start_row+col).value = data
                        xl.cell(column=start_col+1,row=start_row+col).font = Font(color='00000000', bold=True)

                        xl.cell(column=start_col+2,row=start_row+col).value = xl_dict[i][data]    

                    col +=1
            
                

                
            else: # "IS_1K" == xl_sheet[i]
                # keys = xl_dict[i].keys
                is_1k_err = []
                is_500_err = []
                channel = []
                curr_check = []
                curr_check_set = []
                curr_check_500_set = []
                curr_check_500 = []
                for j in xl_dict[i]:
                    if ("acc_err_per" in j) :
                        if ("1k" in j):
                            is_1k_err.append(j)
                        elif ('500_ohm'):
                            is_500_err.append(j)
                 
                for j in range(len(is_1k_err)):
                    channel.append(int(is_1k_err[j].split("_")[0].replace("is",'')))
                    curr_check.append(int(is_1k_err[j].split("_")[-1].replace("uA",'')))
                for j in range(len(is_500_err)):
                    channel.append(int(is_500_err[j].split("_")[0].replace("is",'')))
                    curr_check_500.append(int(is_500_err[j].split("_")[-1].replace("uA",'')))
                
                channel_no = max(channel)
                # print(curr_check)
                # curr_check_set = [*set(curr_check)]
                [curr_check_set.append(x) for x in curr_check if x not in curr_check_set]
                # curr_check_500_set = [*set(curr_check_500)]
                [curr_check_500_set.append(x) for x in curr_check_500 if x not in curr_check_500_set]
                
                # print(curr_check_set)
                # print(channel_no)
                # print(is_1k_err)
                # print('len',len(is_1k_err))

                for j in range(channel_no+2): # insert blank rows for 1k mode err
                    xl.insert_rows(idx=start_row-1) 
                    # xl.cell(row=start_row+1,column=col).value = is_err_xl[channel_no-j-1]


                
                for j in range(len(curr_check_set)):                    
                    xl.cell(row=start_row,column=start_col+3+j).value = curr_check_set[j]
                    xl.cell(row=start_row,column=start_col+3+j).font = Font(color='00000000', bold=True)
                iter = 0
                j=0
                for k in range(int(len(is_1k_err))):
                    # print(j,' ',k)
                    name = is_1k_err[k].split('_')
                    name.pop() # remove last element from list (i.e 100uA)
                    xl.cell(row=start_row+iter+1,column=start_col+2).value = '_'.join(name)
                    xl.cell(row=start_row+iter+1,column=start_col+3+j).value = xl_dict[i][is_1k_err[(j*channel_no + iter)]]
                    iter +=1
                    if iter == channel_no:
                        iter = 0
                        j+=1

                if  curr_check_500_set:
                    for j in range(channel_no+2): # insert blank rows for 500 mode err
                        xl.insert_rows(idx=start_row) 
                        # xl.cell(row=start_row+1,column=col).value = is_err_xl[channel_no-j-1]


                
                for j in range(len(curr_check_500_set)):                    
                    xl.cell(row=start_row,column=start_col+3+j).value = curr_check_500_set[j]
                    xl.cell(row=start_row,column=start_col+3+j).font = Font(color='00000000', bold=True)
                iter = 0
                j =0
                for k in range(len(is_500_err)):
                    # print(j,' ',k)
                    name = is_500_err[k].split('_')
                    name.pop()
                    xl.cell(row=start_row+iter+1,column=start_col+2).value = '_'.join(name)                    
                    xl.cell(row=start_row+iter+1,column=start_col+3+j).value = xl_dict[i][is_500_err[(j*channel_no + iter)]]
                    iter +=1
                    if iter == channel_no:
                        iter = 0
                        j+=1

            # ic_serial_number = read_reg_data(db.jlink,0,"trim.IC_SERIAL.ic_serial")
            temp = str(T_now.strftime("%m/%d/%y - %H:%M")) + "     IC_No.: "+ str(ic_serial_number)
            # print(temp)
            # print(xl_sheet[i])
            if (is_it_rows):
                xl.cell(row=start_row+1,column=1).value = temp
                xl.cell(row=start_row+1,column=1).font = Font(color='00000000', bold=True)
            else:
                temp = str(T_now.strftime("%m/%d/%y - %H:%M"))
                xl.cell(row=start_row,column=start_col+2).value = temp
                xl.cell(row=start_row+1,column=start_col+2).value = "IC_No.: "+ str(ic_serial_number)
                xl.cell(row=start_row+1,column=start_col+2).font = Font(color='00000000', bold=True)
                
            
            for column_cells in xl.columns:
                    length = max(len(str(cell.value)) for cell in column_cells)
                    xl.column_dimensions[column_cells[0].column_letter].width = (length+2) 
                    


        # ------------------------------ make figure ------------------------------
            if ("VS_ADC" == xl_sheet[i]) :
                
                xl = wb["VS_ADC"]
                fig1 = LineChart(auto_axis=True)
                fig2 = LineChart(auto_axis=True)
                fig3 = LineChart(auto_axis=True)
                x_ideal = Reference(xl, min_col= start_col+2, min_row=start_row-3, max_col = start_col+126, max_row = start_row-3)
                x0_test = Reference(xl, min_col= start_col+2, min_row=start_row+1, max_col = start_col+126, max_row = start_row+1)
                x1_test = Reference(xl, min_col= start_col+127, min_row=start_row+1, max_col =start_col+253, max_row = start_row+1)

                x0_err = Reference(xl, min_col= start_col+2, min_row=start_row-2, max_col = start_col+126, max_row = start_row-2)
                x1_err = Reference(xl, min_col= start_col+127, min_row=start_row-2, max_col =start_col+253, max_row = start_row-2)
                
                series = []
                series = Series(x0_test,title="CH-0")
                fig1.append(series)
                # fig1.add_data(data, titles_from_data=True,from_rows=True)  

                series = Series(x1_test,title="CH-1")
                fig2.append(series)
                
                series = Series(x_ideal,title="Ideal")
                fig1.append(series)
                fig2.append(series)
                line1 = fig1.series[1] # styling ideal values
                # line1.marker.symbol = "triangle"
                # line1.marker.graphicalProperties.solidFill = "FF0000" # Marker filling
                # line1.marker.graphicalProperties.line.solidFill = "FF0000" # Marker outline
                line1.graphicalProperties.line.solidFill = "FF0000"
                line1.graphicalProperties.line.dashStyle = "sysDot"
                line1.graphicalProperties.line.width = 100050 # width in EMUs
        
                fig1.title = "VS Accuracy test CH0"
                fig2.title = "VS Accuracy test CH1"
                fig3.title = "Error"
                fig1.x_axis.title = "Input (V)"
                fig1.y_axis.title = "Test (V)"
                
                fig2.x_axis.title = "Input (V)"
                fig2.y_axis.title = "Test (V)"
                
                
                fig1.x_axis.min = min(vs_adc_x)
                fig1.x_axis.max = max(vs_adc_x)
                fig2.x_axis.min = min(vs_adc_x)
                fig2.x_axis.max = max(vs_adc_x)

                # --- error chart -------
                series = Series(x0_err,title="CH-0")
                fig3.append(series)
                series = Series(x1_err,title="CH-1")
                fig3.append(series)
                line0 = fig3.series[0]
                line0.graphicalProperties.line.solidFill = "FF0000"
                line0.graphicalProperties.line.dashStyle = "sysDot"
                line0.graphicalProperties.line.width = 50050 # width in EMUs
                line1 = fig3.series[1]
                line1.graphicalProperties.line.solidFill = "00FF00"
                line1.graphicalProperties.line.dashStyle = "sysDot"
                line1.graphicalProperties.line.width = 50050 # width in EMUs
                fig3.y_axis.title = " Error (mV) "
                # -- update in sheet

                for i in range(len(xl._charts)):
                    try:
                        del xl._charts[i] # delete existing charts in sheet then add_chart
                    except:
                        pass
                
                
                fig1.set_categories(x_ideal)
                fig2.set_categories(x_ideal)
                fig3.set_categories(x_ideal)

                xl.add_chart(fig1, "C11")
                xl.add_chart(fig2, "G11")
                xl.add_chart(fig3, "E25") # <- take error from row, some format error
                

            if ("IS_1K" == xl_sheet[i]):
                xl = wb["IS_1K"]
                fig1 = LineChart(auto_axis = True)
                fig1.title = "1K Accuracy (error"+'%)'
                fig1.x_axis.title = "Current (uA)"
                fig1.y_axis.title = "err %"
                fig2 = LineChart(auto_axis = True)
                fig2.title = "500Ohm Accuracy (error"+'%)'
                fig2.x_axis.title = "Current (uA)"
                fig2.y_axis.title = "err %"
                x_ideal_1k = Reference(xl, min_col= start_col+3, min_row=start_row+channel_no+2, max_col = start_col+3+len(curr_check_set)-1, max_row = start_row+2+(channel_no))
                x_ideal_500 = Reference(xl, min_col= start_col+3, min_row=start_row, max_col = start_col+3+len(curr_check_500_set)-1, max_row = start_row                                                           )

                series = []
                for j in range(channel_no):
                    x_1k_test = Reference(xl, min_col= start_col+3, min_row=start_row+channel_no+3+j, max_col = start_col+3+len(curr_check_set)-1, max_row = start_row+3+j+(channel_no))
                    name = 'ch_'+str(j+1)
                    series = Series(x_1k_test,title=name)
                    fig1.append(series)

                series = []
                for j in range(channel_no):
                    x_500_test = Reference(xl, min_col= start_col+3, min_row=start_row+1+j, max_col = start_col+3+len(curr_check_500_set)-1, max_row = start_row+1+j)
                    name = 'ch_'+str(j+1)
                    series = Series(x_500_test,title=name)
                    fig2.append(series)

                while len(xl._charts)>>0:
                    del xl._charts[0]

                fig1.set_categories(x_ideal_1k)
                fig2.set_categories(x_ideal_500)
                xl.add_chart(fig1, "C11")
                xl.add_chart(fig2, "H11")

    sleep(0.1)
    print("***-> Results in excel:",excel_file)
    wb.save(excel_file)

    temp = "../"*len(pfd_folder.split("\\"))
    os.chdir(temp)
    
    # global debug_option
    # report = db.get_global("running_report")
    # if report['1']: # send message if urgent error
    #     db.report_me_i_did_something(debug_option['email'])

def save_new_trims_to_xl(comment,part_no_string,val_is_0_but_for_correlation_ATE_0_and_SOFT_1_):
    # turn_off_smu()
    ic_serial_number = read_reg_data(db.jlink,0,"trim.IC_SERIAL.ic_serial")
    comment = comment+" "+part_no_string
    this_dir = os.getcwd()
    pfd_folder = db.get_global("pfd_folder")
    if (pfd_folder not in this_dir):
        temp = this_dir+"\\"+pfd_folder
        os.chdir(temp)
    excel_file = db.get_global("excel_file")
    wb = load_workbook(excel_file)

    xl = wb["Trim_Codes"]
    if val_is_0_but_for_correlation_ATE_0_and_SOFT_1_:
        xle = wb["Trim_error"]
    start_row = 7
    start_col = 3

    all_trim_regs = db.get_global("all_trim_regs")
    T_now = datetime.now()

    if all_trim_regs:
        
        xl.insert_cols(idx=start_col+2) # insert blank col at line = idx
        # ic_serial_number = read_reg_data(db.jlink,0,"trim.IC_SERIAL.ic_serial")
        temp = str(T_now.strftime("%m/%d/%y - %H:%M")) + "     IC_No.: "+ str(ic_serial_number)
        xl.cell(row=start_row-1,column=start_col+2).value = temp
        xl.cell(row=start_row-1,column=start_col+2).font = Font(color='00000000', bold=True)
        xl.cell(row=start_row-4,column=start_col+2).value = 'Trimmed'
        xl.cell(row=start_row-4,column=start_col+2).font = Font(color='00000000', bold=True)
        xl.cell(row=start_row-4,column=start_col+3).value = 'NOT Trimmed/ Reset'
        xl.cell(row=start_row-4,column=start_col+3).font = Font(color='FF0000', bold=True)

        xl.cell(row=start_row-2,column=start_col+2).value = comment
        xl.cell(row=start_row-2,column=start_col+2).font = Font(color='00000000', bold=True)
        xl.cell(row=start_row-2,column=start_col+1).value = "Error from previous"

        if val_is_0_but_for_correlation_ATE_0_and_SOFT_1_:
            xle.insert_cols(idx=start_col+1) # insert blank col at line = idx
            xle.cell(row=start_row-2,column=start_col+1).value = part_no_string
            xle.cell(row=start_row-2,column=start_col+1).font = Font(color='00000000', bold=True)
            xle.cell(row=start_row-4,column=start_col).value = 'Error between ATE and Soft trims'
            xle.cell(row=start_row-4,column=start_col).font = Font(color='FF0000', bold=True)
            # xle.cell(row=start_row-4,column=start_col+3).value = ''

        
        index = 0  
        for i in all_trim_regs:
            xl.cell(row=start_row+index,column=start_col).value = i #name of registers
            if val_is_0_but_for_correlation_ATE_0_and_SOFT_1_:
                xle.cell(row=start_row+index,column=start_col).value = i
            try: # for names of "is" trims
                xl.cell(row=start_row+index,column=start_col-1).value = all_trim_regs[i][2]
                if val_is_0_but_for_correlation_ATE_0_and_SOFT_1_:
                    xle.cell(row=start_row+index,column=start_col-1).value = all_trim_regs[i][2]
            except:
                pass
            xl.cell(row=start_row+index,column=start_col+2).value = all_trim_regs[i][1]
            if all_trim_regs[i][0] == all_trim_regs[i][1]:  # check with default
                xl.cell(row=start_row+index,column=start_col+2).font = Font(color='FF0000', bold=False)
            index +=1
        xl.cell(row=start_row-4,column=start_col+4).value = ''



    ##### comparing trims between saves, TRIM CORRELATION for IS, as the trims are not mentioned in the regstermap.json
    try:
        names = []
        old_data = []
        new_data = []
        
        old_trim = ''
        for i in range(len(all_trim_regs)):
            names.append(xl.cell(row=start_row+i,column = start_col).value)
            old_data.append(xl.cell(row=start_row+i,column = start_col+3).value)
            new_data.append(xl.cell(row=start_row+i,column = start_col+2).value)

        for i in range(len(all_trim_regs)):
            difference = []
            if '0x' in names[i]:
                old_trim = old_data[i]
                old_trim = old_data[i].split('(')
                old_trim = old_trim[1].replace(')','').split(' ')
                while '' in old_trim:
                    old_trim.remove('')
                new_trim = new_data[i]
                new_trim = new_data[i].split('(')
                new_trim = new_trim[1].replace(')','').split(' ')
                while '' in new_trim:
                    new_trim.remove('')
                for j in range(len(new_trim)):
                    difference.append(str(int(new_trim[j])-int(old_trim[j])))
                difference = ' '.join(difference)
                
            else:
                difference = str(int(new_data[i])-int(old_data[i]))

            xl.cell(row=start_row+i,column=start_col+1).value = difference
            xl.cell(row=start_row+i,column=start_col+1).font = Font(color='0000F0', bold=False)

            if val_is_0_but_for_correlation_ATE_0_and_SOFT_1_:
                xle.cell(row=start_row+i,column=start_col+1).value = difference


        for column_cells in xl.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            xl.column_dimensions[column_cells[0].column_letter].width = (length+2)
    
        if val_is_0_but_for_correlation_ATE_0_and_SOFT_1_:
            for column_cells in xle.columns:
                length = max(len(str(cell.value)) for cell in column_cells)
                xle.column_dimensions[column_cells[0].column_letter].width = (length+2)

    except:
        pass    

    wb.save(excel_file)
    temp = "../"*len(pfd_folder.split("\\"))
    os.chdir(temp)



